"""
Auto-format Excel columns with width adjustment and styling.
"""

from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


def auto_format_excel_columns(
    file_path: str | Path,
    *,
    sheet_name: str | None = None,
    auto_width: bool = True,
    freeze_header: bool = True,
    bold_header: bool = True,
    header_fill_color: str | None = "CCCCCC",
    align_text: str = "left",
) -> None:
    """
    Auto-format Excel sheet with column width, header styling, and alignment.

    Applies professional formatting including auto-sizing columns, freezing
    header row, bold headers, and text alignment. Modifies file in place.

    Parameters
    ----------
    file_path : str | Path
        Path to Excel file (.xlsx) to format.
    sheet_name : str | None, optional
        Name of sheet to format (by default None for active sheet).
    auto_width : bool, optional
        Auto-size column widths based on content (by default True).
    freeze_header : bool, optional
        Freeze first row as header (by default True).
    bold_header : bool, optional
        Make header row bold (by default True).
    header_fill_color : str | None, optional
        Hex color for header background (by default "CCCCCC" gray).
    align_text : str, optional
        Text alignment: 'left', 'center', 'right' (by default 'left').

    Returns
    -------
    None
        File is modified in place.

    Raises
    ------
    TypeError
        If parameters are of wrong type.
    FileNotFoundError
        If file doesn't exist.
    ValueError
        If parameters have invalid values.

    Examples
    --------
    >>> # Basic auto-formatting
    >>> auto_format_excel_columns('data.xlsx')

    >>> # Custom formatting
    >>> auto_format_excel_columns(
    ...     'data.xlsx',
    ...     sheet_name='Sales',
    ...     header_fill_color='4472C4',
    ...     align_text='center'
    ... )

    >>> # Minimal formatting
    >>> auto_format_excel_columns(
    ...     'data.xlsx',
    ...     freeze_header=False,
    ...     bold_header=False,
    ...     header_fill_color=None
    ... )

    Notes
    -----
    Auto-width calculation considers cell content length with padding.
    Maximum column width is capped at 50 characters for readability.

    Complexity
    ----------
    Time: O(n*m) where n=rows, m=columns
    Space: O(1)
    """
    # Type validation
    if not isinstance(file_path, (str, Path)):
        raise TypeError(
            f"file_path must be str or Path, got {type(file_path).__name__}"
        )
    if sheet_name is not None and not isinstance(sheet_name, str):
        raise TypeError(
            f"sheet_name must be str or None, got {type(sheet_name).__name__}"
        )
    if not isinstance(auto_width, bool):
        raise TypeError(f"auto_width must be bool, got {type(auto_width).__name__}")
    if not isinstance(freeze_header, bool):
        raise TypeError(
            f"freeze_header must be bool, got {type(freeze_header).__name__}"
        )
    if not isinstance(bold_header, bool):
        raise TypeError(f"bold_header must be bool, got {type(bold_header).__name__}")
    if header_fill_color is not None and not isinstance(header_fill_color, str):
        raise TypeError(
            f"header_fill_color must be str or None, got {type(header_fill_color).__name__}"
        )
    if not isinstance(align_text, str):
        raise TypeError(f"align_text must be str, got {type(align_text).__name__}")

    # Value validation
    if not Path(file_path).exists():
        raise FileNotFoundError(f"Excel file not found: {file_path}")
    if align_text not in ("left", "center", "right"):
        raise ValueError(
            f"align_text must be 'left', 'center', or 'right', got '{align_text}'"
        )

    # Load workbook
    wb = openpyxl.load_workbook(file_path)

    # Get worksheet
    if sheet_name is None:
        ws = wb.active
        if ws is None:
            raise ValueError("Workbook has no active sheet")
    else:
        if sheet_name not in wb.sheetnames:
            raise ValueError(f"Sheet '{sheet_name}' not found in workbook")
        ws = wb[sheet_name]

    # Freeze header row
    if freeze_header:
        ws.freeze_panes = "A2"

    # Calculate column widths and format
    if auto_width or bold_header or header_fill_color or align_text != "left":
        # Get dimensions
        max_row = ws.max_row
        max_col = ws.max_column

        if max_row > 0 and max_col > 0:
            # Auto-size columns
            if auto_width:
                for col_idx in range(1, max_col + 1):
                    col_letter = get_column_letter(col_idx)
                    max_length = 0

                    for row_idx in range(1, max_row + 1):
                        cell = ws.cell(row=row_idx, column=col_idx)
                        if cell.value is not None:
                            cell_length = len(str(cell.value))
                            max_length = max(max_length, cell_length)

                    # Add padding and cap at 50
                    adjusted_width = min(max_length + 2, 50)
                    ws.column_dimensions[col_letter].width = adjusted_width

            # Format header row
            if max_row >= 1:
                for col_idx in range(1, max_col + 1):
                    cell = ws.cell(row=1, column=col_idx)

                    if bold_header:
                        cell.font = Font(bold=True)

                    if header_fill_color:
                        cell.fill = PatternFill(
                            start_color=header_fill_color,
                            end_color=header_fill_color,
                            fill_type="solid",
                        )

            # Apply text alignment to all cells
            if align_text != "left":
                for row in ws.iter_rows(
                    min_row=1, max_row=max_row, min_col=1, max_col=max_col
                ):
                    for cell in row:
                        cell.alignment = Alignment(horizontal=align_text)

    # Save workbook
    wb.save(file_path)


__all__ = ["auto_format_excel_columns"]

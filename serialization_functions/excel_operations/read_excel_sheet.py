"""
Read data from Excel sheet.
"""

from typing import Any

from openpyxl import load_workbook


def read_excel_sheet(
    file_path: str,
    sheet_name: str | int = 0,
    min_row: int | None = None,
    max_row: int | None = None,
    min_col: int | None = None,
    max_col: int | None = None,
    values_only: bool = True,
) -> list[list[Any]]:
    """
    Read data from an Excel sheet with range support.

    Parameters
    ----------
    file_path : str
        Path to Excel file (.xlsx).
    sheet_name : str | int, optional
        Sheet name or index (by default 0).
    min_row : int | None, optional
        Starting row (1-indexed) (by default None).
    max_row : int | None, optional
        Ending row (1-indexed) (by default None).
    min_col : int | None, optional
        Starting column (1-indexed) (by default None).
    max_col : int | None, optional
        Ending column (1-indexed) (by default None).
    values_only : bool, optional
        Return cell values only (by default True).

    Returns
    -------
    list[list[Any]]
        2D list of cell values.

    Raises
    ------
    TypeError
        If parameters are of wrong type.
    ValueError
        If row/column indices are invalid.
    FileNotFoundError
        If file doesn't exist.
    ImportError
        If openpyxl is not installed.

    Examples
    --------
    >>> data = read_excel_sheet('data.xlsx')
    >>> data[0]
    ['Name', 'Age', 'City']

    >>> data = read_excel_sheet('data.xlsx', sheet_name='Sheet2', min_row=2, max_row=10)

    >>> data = read_excel_sheet('data.xlsx', min_col=1, max_col=3, values_only=True)

    Notes
    -----
    Requires openpyxl package.
    Row and column indices are 1-based (Excel convention).

    Complexity
    ----------
    Time: O(n*m), Space: O(n*m), where n is rows, m is columns
    """
    if not isinstance(file_path, str):
        raise TypeError(f"file_path must be a string, got {type(file_path).__name__}")
    
    if not isinstance(sheet_name, (str, int)):
        raise TypeError(f"sheet_name must be str or int, got {type(sheet_name).__name__}")
    
    if not isinstance(values_only, bool):
        raise TypeError(f"values_only must be a boolean, got {type(values_only).__name__}")
    
    # Validate row/column indices
    for param_name, param_value in [
        ("min_row", min_row), ("max_row", max_row),
        ("min_col", min_col), ("max_col", max_col)
    ]:
        if param_value is not None:
            if not isinstance(param_value, int):
                raise TypeError(f"{param_name} must be int or None, got {type(param_value).__name__}")
            if param_value < 1:
                raise ValueError(f"{param_name} must be >= 1, got {param_value}")
    
    # Load workbook and sheet
    wb = load_workbook(file_path, read_only=True, data_only=values_only)
    
    if isinstance(sheet_name, int):
        if sheet_name < 0 or sheet_name >= len(wb.sheetnames):
            raise ValueError(f"sheet_name index {sheet_name} out of range (0-{len(wb.sheetnames)-1})")
        ws = wb.worksheets[sheet_name]
    else:
        if sheet_name not in wb.sheetnames:
            raise ValueError(f"sheet_name '{sheet_name}' not found in workbook")
        ws = wb[sheet_name]
    
    # Read data
    data = []
    for row in ws.iter_rows(
        min_row=min_row,
        max_row=max_row,
        min_col=min_col,
        max_col=max_col,
        values_only=values_only
    ):
        data.append(list(row))
    
    wb.close()
    return data


__all__ = ['read_excel_sheet']

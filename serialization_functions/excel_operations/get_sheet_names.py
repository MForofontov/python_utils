"""
Get list of sheet names from Excel workbook.
"""

from openpyxl import load_workbook


def get_sheet_names(file_path: str) -> list[str]:
    """
    Get list of all sheet names in an Excel workbook.

    Parameters
    ----------
    file_path : str
        Path to Excel file (.xlsx).

    Returns
    -------
    list[str]
        List of sheet names.

    Raises
    ------
    TypeError
        If file_path is not a string.
    FileNotFoundError
        If file doesn't exist.
    ImportError
        If openpyxl is not installed.

    Examples
    --------
    >>> sheets = get_sheet_names('data.xlsx')
    >>> sheets
    ['Sheet1', 'Sheet2', 'Summary']

    >>> for sheet in get_sheet_names('workbook.xlsx'):
    ...     print(f"Processing {sheet}")

    Notes
    -----
    Requires openpyxl package.
    Returns sheets in order they appear in workbook.

    Complexity
    ----------
    Time: O(n), Space: O(n), where n is number of sheets
    """
    if not isinstance(file_path, str):
        raise TypeError(f"file_path must be a string, got {type(file_path).__name__}")
    
    wb = load_workbook(file_path, read_only=True)
    sheet_names = wb.sheetnames
    wb.close()
    
    return sheet_names  # type: ignore[no-any-return]


__all__ = ['get_sheet_names']

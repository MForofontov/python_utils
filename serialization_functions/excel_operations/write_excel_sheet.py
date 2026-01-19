"""
Write data to Excel sheet.
"""

from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook


def write_excel_sheet(
    data: list[list[Any]],
    file_path: str,
    sheet_name: str = "Sheet1",
    mode: str = "w",
    start_row: int = 1,
    start_col: int = 1,
) -> None:
    """
    Write 2D data to an Excel sheet.

    Parameters
    ----------
    data : list[list[Any]]
        2D list of values to write.
    file_path : str
        Path to Excel file (.xlsx).
    sheet_name : str, optional
        Sheet name (by default "Sheet1").
    mode : str, optional
        'w' to overwrite, 'a' to append (by default "w").
    start_row : int, optional
        Starting row (1-indexed) (by default 1).
    start_col : int, optional
        Starting column (1-indexed) (by default 1).

    Returns
    -------
    None

    Raises
    ------
    TypeError
        If parameters are of wrong type.
    ValueError
        If data is empty or indices invalid.
    ImportError
        If openpyxl is not installed.

    Examples
    --------
    >>> data = [['Name', 'Age'], ['Alice', 30], ['Bob', 25]]
    >>> write_excel_sheet(data, 'output.xlsx')

    >>> write_excel_sheet(data, 'output.xlsx', sheet_name='People', start_row=2)

    >>> write_excel_sheet(data, 'existing.xlsx', mode='a')

    Notes
    -----
    Requires openpyxl package.
    Creates parent directories if they don't exist.

    Complexity
    ----------
    Time: O(n*m), Space: O(n*m), where n is rows, m is columns
    """
    if not isinstance(data, list):
        raise TypeError(f"data must be a list, got {type(data).__name__}")

    if not data:
        raise ValueError("data cannot be empty")

    if not all(isinstance(row, list) for row in data):
        raise TypeError("all elements in data must be lists")

    if not isinstance(file_path, str):
        raise TypeError(f"file_path must be a string, got {type(file_path).__name__}")

    if not isinstance(sheet_name, str):
        raise TypeError(f"sheet_name must be a string, got {type(sheet_name).__name__}")

    if not isinstance(mode, str):
        raise TypeError(f"mode must be a string, got {type(mode).__name__}")

    if mode not in ("w", "a"):
        raise ValueError(f"mode must be 'w' or 'a', got '{mode}'")

    if not isinstance(start_row, int):
        raise TypeError(f"start_row must be an integer, got {type(start_row).__name__}")

    if not isinstance(start_col, int):
        raise TypeError(f"start_col must be an integer, got {type(start_col).__name__}")

    if start_row < 1:
        raise ValueError(f"start_row must be >= 1, got {start_row}")

    if start_col < 1:
        raise ValueError(f"start_col must be >= 1, got {start_col}")

    # Create parent directories
    Path(file_path).parent.mkdir(parents=True, exist_ok=True)

    # Load or create workbook
    if mode == "a" and Path(file_path).exists():
        wb = load_workbook(file_path)
        # If sheet_name is default "Sheet1" and it doesn't exist, use active sheet
        if sheet_name == "Sheet1" and sheet_name not in wb.sheetnames:
            ws = wb.active
        elif sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
        else:
            ws = wb.create_sheet(sheet_name)
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name

    # Write data
    for row_idx, row_data in enumerate(data, start=start_row):
        for col_idx, value in enumerate(row_data, start=start_col):
            ws.cell(row=row_idx, column=col_idx, value=value)

    wb.save(file_path)
    wb.close()


__all__ = ["write_excel_sheet"]

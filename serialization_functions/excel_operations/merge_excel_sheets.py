"""
Merge multiple Excel sheets into a single sheet.
"""

from pathlib import Path
from typing import Any
from collections.abc import Sequence

import openpyxl
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet


def merge_excel_sheets(
    input_file: str | Path,
    output_file: str | Path,
    *,
    sheet_names: Sequence[str] | None = None,
    output_sheet_name: str = "Merged",
    include_headers: bool = True,
    skip_duplicates: bool = False,
) -> int:
    """
    Merge multiple sheets from an Excel workbook into a single sheet.

    Combines data from multiple sheets with consistent structure. Handles
    header rows and duplicate detection across sheets.

    Parameters
    ----------
    input_file : str | Path
        Path to input Excel file (.xlsx).
    output_file : str | Path
        Path to output Excel file (.xlsx).
    sheet_names : Sequence[str] | None, optional
        Names of sheets to merge (by default None for all).
    output_sheet_name : str, optional
        Name of output merged sheet (by default "Merged").
    include_headers : bool, optional
        Whether sheets have header rows (by default True).
    skip_duplicates : bool, optional
        Skip duplicate rows across sheets (by default False).

    Returns
    -------
    int
        Number of data rows written to output.

    Raises
    ------
    TypeError
        If parameters are of wrong type.
    FileNotFoundError
        If input file doesn't exist.
    ValueError
        If sheet names are invalid or headers don't match.

    Examples
    --------
    >>> # Merge all sheets
    >>> rows = merge_excel_sheets('data.xlsx', 'merged.xlsx')
    >>> rows
    1500

    >>> # Merge specific sheets
    >>> merge_excel_sheets(
    ...     'data.xlsx',
    ...     'merged.xlsx',
    ...     sheet_names=['Sheet1', 'Sheet2'],
    ...     skip_duplicates=True
    ... )
    1234

    >>> # Merge without headers
    >>> merge_excel_sheets(
    ...     'data.xlsx',
    ...     'merged.xlsx',
    ...     include_headers=False
    ... )
    1550

    Notes
    -----
    All merged sheets must have the same column structure if include_headers=True.
    First sheet's header is used; subsequent sheets are validated against it.

    Complexity
    ----------
    Time: O(n*m) where n=total rows, m=columns per row
    Space: O(k) where k=unique rows if skip_duplicates=True
    """
    # Type validation
    if not isinstance(input_file, (str, Path)):
        raise TypeError(
            f"input_file must be str or Path, got {type(input_file).__name__}"
        )
    if not isinstance(output_file, (str, Path)):
        raise TypeError(
            f"output_file must be str or Path, got {type(output_file).__name__}"
        )
    if sheet_names is not None and not isinstance(sheet_names, (list, tuple)):
        raise TypeError(
            f"sheet_names must be sequence or None, got {type(sheet_names).__name__}"
        )
    if not isinstance(output_sheet_name, str):
        raise TypeError(
            f"output_sheet_name must be str, got {type(output_sheet_name).__name__}"
        )
    if not isinstance(include_headers, bool):
        raise TypeError(
            f"include_headers must be bool, got {type(include_headers).__name__}"
        )
    if not isinstance(skip_duplicates, bool):
        raise TypeError(
            f"skip_duplicates must be bool, got {type(skip_duplicates).__name__}"
        )

    # Value validation
    if not Path(input_file).exists():
        raise FileNotFoundError(f"Input file not found: {input_file}")
    if not output_sheet_name:
        raise ValueError("output_sheet_name cannot be empty")

    # Load workbook
    wb = openpyxl.load_workbook(input_file, data_only=True)

    # Determine sheets to merge
    if sheet_names is None:
        sheets_to_merge = wb.sheetnames
    else:
        sheets_to_merge = list(sheet_names)
        # Validate sheet names exist
        for name in sheets_to_merge:
            if name not in wb.sheetnames:
                raise ValueError(f"Sheet '{name}' not found in workbook")

    if len(sheets_to_merge) == 0:
        raise ValueError("No sheets to merge")

    # Create output workbook
    output_wb = Workbook()
    output_ws = output_wb.active
    if output_ws is None:
        output_ws = output_wb.create_sheet(output_sheet_name)
    else:
        output_ws.title = output_sheet_name

    header: tuple[Any, ...] | None = None
    rows_written = 0
    seen_rows: set[tuple[Any, ...]] = set() if skip_duplicates else set()

    for idx, sheet_name in enumerate(sheets_to_merge):
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))

        if len(rows) == 0:
            continue

        start_idx = 0

        # Handle header
        if include_headers:
            if len(rows) < 1:
                continue

            sheet_header = rows[0]

            if idx == 0:
                # First sheet - establish header
                header = sheet_header
                output_ws.append(header)
            else:
                # Validate header matches
                if sheet_header != header:
                    raise ValueError(
                        f"Header mismatch in sheet '{sheet_name}'. "
                        f"Expected {header}, got {sheet_header}"
                    )

            start_idx = 1  # Skip header row

        # Write data rows
        for row in rows[start_idx:]:
            # Skip empty rows
            if not any(cell is not None for cell in row):
                continue

            if skip_duplicates:
                if row in seen_rows:
                    continue
                seen_rows.add(row)

            output_ws.append(row)
            rows_written += 1

    # Save output workbook
    output_wb.save(output_file)

    return rows_written


__all__ = ['merge_excel_sheets']

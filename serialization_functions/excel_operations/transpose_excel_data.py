"""
Transpose Excel data (rows to columns).
"""

from pathlib import Path

import openpyxl
from openpyxl.workbook import Workbook


def transpose_excel_data(
    input_file: str | Path,
    output_file: str | Path,
    *,
    sheet_name: str | None = None,
    output_sheet_name: str = "Transposed",
) -> tuple[int, int]:
    """
    Transpose Excel sheet data (convert rows to columns).

    Reads data from input sheet and writes transposed version to output.
    Useful for pivoting data or converting between row/column-oriented formats.

    Parameters
    ----------
    input_file : str | Path
        Path to input Excel file (.xlsx).
    output_file : str | Path
        Path to output Excel file (.xlsx).
    sheet_name : str | None, optional
        Name of sheet to transpose (by default None for active sheet).
    output_sheet_name : str, optional
        Name of output transposed sheet (by default "Transposed").

    Returns
    -------
    tuple[int, int]
        (original_rows, original_cols) - dimensions of source data.

    Raises
    ------
    TypeError
        If parameters are of wrong type.
    FileNotFoundError
        If input file doesn't exist.
    ValueError
        If sheet name is invalid.

    Examples
    --------
    >>> # Transpose entire sheet
    >>> rows, cols = transpose_excel_data('data.xlsx', 'transposed.xlsx')
    >>> rows, cols
    (100, 5)

    >>> # Transpose specific sheet
    >>> transpose_excel_data(
    ...     'data.xlsx',
    ...     'transposed.xlsx',
    ...     sheet_name='Data',
    ...     output_sheet_name='Pivoted'
    ... )
    (100, 5)

    Notes
    -----
    Empty cells are preserved in transposed output.
    For a 100x5 sheet, output will be 5x100.

    Complexity
    ----------
    Time: O(n*m) where n=rows, m=columns
    Space: O(n*m) to store all data
    """
    # Type validation
    if not isinstance(input_file, (str, Path)):
        raise TypeError(
            f"input_file must be str or Path, got {type(input_file).__name__}"
        )
    if not isinstance(output_file, (str, Path)):
        raise TypeError(
            f"output_file must be str or Path, got {type(output_file).__name__}"
        )
    if sheet_name is not None and not isinstance(sheet_name, str):
        raise TypeError(
            f"sheet_name must be str or None, got {type(sheet_name).__name__}"
        )
    if not isinstance(output_sheet_name, str):
        raise TypeError(
            f"output_sheet_name must be str, got {type(output_sheet_name).__name__}"
        )

    # Value validation
    if not Path(input_file).exists():
        raise FileNotFoundError(f"Input file not found: {input_file}")
    if not output_sheet_name:
        raise ValueError("output_sheet_name cannot be empty")

    # Load workbook
    wb = openpyxl.load_workbook(input_file, data_only=True)

    # Get worksheet
    if sheet_name is None:
        ws = wb.active
        if ws is None:
            raise ValueError("Workbook has no active sheet")
    else:
        if sheet_name not in wb.sheetnames:
            raise ValueError(f"Sheet '{sheet_name}' not found in workbook")
        ws = wb[sheet_name]

    # Read all data
    data = list(ws.iter_rows(values_only=True))

    if len(data) == 0:
        # Empty sheet
        original_rows = 0
        original_cols = 0
    else:
        original_rows = len(data)
        original_cols = max(len(row) for row in data) if data else 0

    # Transpose data
    # Handle irregular row lengths by padding with None
    max_cols = original_cols
    padded_data = [
        list(row) + [None] * (max_cols - len(row))
        for row in data
    ]

    transposed_data = list(zip(*padded_data)) if padded_data else []

    # Create output workbook
    output_wb = Workbook()
    output_ws = output_wb.active
    if output_ws is None:
        output_ws = output_wb.create_sheet(output_sheet_name)
    else:
        output_ws.title = output_sheet_name

    # Write transposed data
    for row in transposed_data:
        output_ws.append(row)

    # Save output workbook
    output_wb.save(output_file)

    return original_rows, original_cols


__all__ = ['transpose_excel_data']

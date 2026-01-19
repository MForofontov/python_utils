"""
Write data to specific Excel range.
"""

from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.utils import column_index_from_string
from openpyxl.utils.cell import coordinate_from_string


def write_excel_range(
    data: list[list[Any]],
    file_path: str,
    sheet_name: str = "Sheet1",
    start_cell: str = "A1",
    mode: str = "w",
) -> None:
    """
    Write 2D data to a specific Excel range starting at given cell.

    Parameters
    ----------
    data : list[list[Any]]
        2D list of values to write.
    file_path : str
        Path to Excel file (.xlsx).
    sheet_name : str, optional
        Sheet name (by default "Sheet1").
    start_cell : str, optional
        Starting cell in Excel notation (by default "A1").
    mode : str, optional
        'w' to overwrite, 'a' to append (by default "w").

    Returns
    -------
    None

    Raises
    ------
    TypeError
        If parameters are of wrong type.
    ValueError
        If data is empty or start_cell invalid.
    ImportError
        If openpyxl is not installed.

    Examples
    --------
    >>> data = [['Name', 'Age'], ['Alice', 30]]
    >>> write_excel_range(data, 'output.xlsx', start_cell='B2')

    >>> write_excel_range(data, 'existing.xlsx', sheet_name='People', mode='a')

    Notes
    -----
    Requires openpyxl package.
    Creates parent directories if they don't exist.

    Complexity
    ----------
    Time: O(n*m), Space: O(n*m), where n is rows, m is columns
    """
    if not isinstance(data, list):
        raise TypeError(f"data must be a list, got {type(data).__name__}")

    if not data:
        raise ValueError("data cannot be empty")

    if not all(isinstance(row, list) for row in data):
        raise TypeError("all elements in data must be lists")

    if not isinstance(file_path, str):
        raise TypeError(f"file_path must be a string, got {type(file_path).__name__}")

    if not isinstance(sheet_name, str):
        raise TypeError(f"sheet_name must be a string, got {type(sheet_name).__name__}")

    if not isinstance(start_cell, str):
        raise TypeError(f"start_cell must be a string, got {type(start_cell).__name__}")

    if not isinstance(mode, str):
        raise TypeError(f"mode must be a string, got {type(mode).__name__}")

    if mode not in ("w", "a"):
        raise ValueError(f"mode must be 'w' or 'a', got '{mode}'")

    # Parse start cell
    try:
        col_letter, row_num = coordinate_from_string(start_cell)
        start_col = column_index_from_string(col_letter)
        start_row = row_num
    except Exception as e:
        raise ValueError(f"Invalid start_cell format '{start_cell}'") from e

    # Create parent directories
    Path(file_path).parent.mkdir(parents=True, exist_ok=True)

    # Load or create workbook
    if mode == "a" and Path(file_path).exists():
        wb = load_workbook(file_path)
        # If sheet_name is default "Sheet1" and it doesn't exist, use active sheet
        if sheet_name == "Sheet1" and sheet_name not in wb.sheetnames:
            ws = wb.active
        elif sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
        else:
            ws = wb.create_sheet(sheet_name)
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name

    # Write data
    for row_idx, row_data in enumerate(data):
        for col_idx, value in enumerate(row_data):
            ws.cell(row=start_row + row_idx, column=start_col + col_idx, value=value)

    wb.save(file_path)
    wb.close()


__all__ = ["write_excel_range"]

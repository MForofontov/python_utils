"""
Read specific cell range from Excel sheet.
"""

from typing import Any


def read_excel_range(
    file_path: str,
    sheet_name: str | int = 0,
    range_string: str | None = None,
    values_only: bool = True,
) -> list[list[Any]]:
    """
    Read a specific range from Excel sheet using Excel notation.

    Parameters
    ----------
    file_path : str
        Path to Excel file (.xlsx).
    sheet_name : str | int, optional
        Sheet name or index (by default 0).
    range_string : str | None, optional
        Excel range notation like 'A1:C10' (by default None for all).
    values_only : bool, optional
        Return cell values only (by default True).

    Returns
    -------
    list[list[Any]]
        2D list of cell values.

    Raises
    ------
    TypeError
        If parameters are of wrong type.
    ValueError
        If range_string is invalid.
    FileNotFoundError
        If file doesn't exist.
    ImportError
        If openpyxl is not installed.

    Examples
    --------
    >>> data = read_excel_range('data.xlsx', range_string='A1:C10')
    >>> len(data)
    10

    >>> data = read_excel_range('data.xlsx', sheet_name='Sheet2', range_string='B2:D5')

    >>> data = read_excel_range('data.xlsx')  # Read entire sheet

    Notes
    -----
    Requires openpyxl package.
    Uses Excel range notation (e.g., 'A1:B10').

    Complexity
    ----------
    Time: O(n*m), Space: O(n*m), where n is rows, m is columns
    """
    try:
        from openpyxl import load_workbook
    except ImportError as e:
        raise ImportError("openpyxl is required. Install with: pip install openpyxl") from e
    
    if not isinstance(file_path, str):
        raise TypeError(f"file_path must be a string, got {type(file_path).__name__}")
    
    if not isinstance(sheet_name, (str, int)):
        raise TypeError(f"sheet_name must be str or int, got {type(sheet_name).__name__}")
    
    if range_string is not None and not isinstance(range_string, str):
        raise TypeError(f"range_string must be str or None, got {type(range_string).__name__}")
    
    if not isinstance(values_only, bool):
        raise TypeError(f"values_only must be a boolean, got {type(values_only).__name__}")
    
    # Load workbook and sheet
    wb = load_workbook(file_path, read_only=True, data_only=values_only)
    
    if isinstance(sheet_name, int):
        if sheet_name < 0 or sheet_name >= len(wb.sheetnames):
            raise ValueError(f"sheet_name index {sheet_name} out of range")
        ws = wb.worksheets[sheet_name]
    else:
        if sheet_name not in wb.sheetnames:
            raise ValueError(f"sheet_name '{sheet_name}' not found")
        ws = wb[sheet_name]
    
    # Read range
    if range_string:
        cell_range = ws[range_string]
        # Handle single cell
        if not isinstance(cell_range, tuple):
            data = [[cell_range.value if values_only else cell_range]]
        # Handle single row
        elif not isinstance(cell_range[0], tuple):
            data = [[cell.value if values_only else cell for cell in cell_range]]
        # Handle 2D range
        else:
            data = [[cell.value if values_only else cell for cell in row] for row in cell_range]
    else:
        data = [[cell.value if values_only else cell for cell in row] for row in ws.iter_rows()]
    
    wb.close()
    return data


__all__ = ['read_excel_range']

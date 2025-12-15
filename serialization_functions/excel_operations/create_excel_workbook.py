"""
Create new Excel workbook with multiple sheets.
"""

from typing import Any
from pathlib import Path


def create_excel_workbook(
    file_path: str,
    sheets: dict[str, list[list[Any]]] | None = None,
) -> None:
    """
    Create a new Excel workbook with multiple sheets.

    Parameters
    ----------
    file_path : str
        Path to output Excel file (.xlsx).
    sheets : dict[str, list[list[Any]]] | None, optional
        Dictionary mapping sheet names to 2D data (by default None).

    Returns
    -------
    None

    Raises
    ------
    TypeError
        If parameters are of wrong type.
    ValueError
        If sheet names or data are invalid.
    ImportError
        If openpyxl is not installed.

    Examples
    --------
    >>> data = {
    ...     'People': [['Name', 'Age'], ['Alice', 30]],
    ...     'Cities': [['City', 'Country'], ['NYC', 'USA']]
    ... }
    >>> create_excel_workbook('output.xlsx', sheets=data)

    >>> create_excel_workbook('empty.xlsx')

    Notes
    -----
    Requires openpyxl package.
    Creates parent directories if they don't exist.
    If sheets is None, creates workbook with one empty sheet.

    Complexity
    ----------
    Time: O(n*m*s), Space: O(n*m*s), where s is sheets, n is rows, m is columns
    """
    try:
        from openpyxl import Workbook
    except ImportError as e:
        raise ImportError("openpyxl is required. Install with: pip install openpyxl") from e
    
    if not isinstance(file_path, str):
        raise TypeError(f"file_path must be a string, got {type(file_path).__name__}")
    
    if sheets is not None and not isinstance(sheets, dict):
        raise TypeError(f"sheets must be a dict or None, got {type(sheets).__name__}")
    
    # Create parent directories
    Path(file_path).parent.mkdir(parents=True, exist_ok=True)
    
    wb = Workbook()
    
    if sheets:
        # Remove default sheet
        if 'Sheet' in wb.sheetnames:
            wb.remove(wb['Sheet'])
        
        for sheet_name, data in sheets.items():
            if not isinstance(sheet_name, str):
                raise TypeError(f"sheet name must be a string, got {type(sheet_name).__name__}")
            
            if not isinstance(data, list):
                raise TypeError(f"sheet data must be a list, got {type(data).__name__}")
            
            if not all(isinstance(row, list) for row in data):
                raise TypeError("all rows in sheet data must be lists")
            
            ws = wb.create_sheet(sheet_name)
            
            for row_data in data:
                ws.append(row_data)
    
    wb.save(file_path)
    wb.close()


__all__ = ['create_excel_workbook']

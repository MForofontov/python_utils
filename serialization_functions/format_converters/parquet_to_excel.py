"""
Convert Parquet to Excel with formatting.
"""

from pathlib import Path

import openpyxl
import pyarrow.parquet as pq
from openpyxl.styles import Font, PatternFill
from openpyxl.workbook import Workbook


def parquet_to_excel(
    input_file: str | Path,
    output_file: str | Path,
    *,
    sheet_name: str = "Data",
    columns: list[str] | None = None,
    max_rows: int | None = None,
    auto_format: bool = True,
) -> int:
    """
    Convert Parquet file to Excel format with optional formatting.

    Exports Parquet data to Excel with professional formatting including
    auto-sized columns, bold headers, and frozen header row.

    Parameters
    ----------
    input_file : str | Path
        Path to input Parquet file.
    output_file : str | Path
        Path to output Excel file (.xlsx).
    sheet_name : str, optional
        Name of Excel sheet (by default "Data").
    columns : list[str] | None, optional
        Columns to export (by default None for all).
    max_rows : int | None, optional
        Maximum rows to export (by default None for all).
    auto_format : bool, optional
        Apply auto-formatting (by default True).

    Returns
    -------
    int
        Number of data rows written to Excel.

    Raises
    ------
    TypeError
        If parameters are of wrong type.
    FileNotFoundError
        If input file doesn't exist.
    ValueError
        If parameters have invalid values.

    Examples
    --------
    >>> # Basic conversion
    >>> rows = parquet_to_excel('data.parquet', 'data.xlsx')
    >>> rows
    10000

    >>> # Export specific columns
    >>> parquet_to_excel(
    ...     'data.parquet',
    ...     'data.xlsx',
    ...     columns=['id', 'name', 'email']
    ... )
    10000

    >>> # Limit rows
    >>> parquet_to_excel(
    ...     'data.parquet',
    ...     'preview.xlsx',
    ...     max_rows=100,
    ...     auto_format=True
    ... )
    100

    Notes
    -----
    Excel has a limit of 1,048,576 rows. Use max_rows to avoid exceeding limit.
    Auto-formatting includes: bold headers, frozen top row, auto-sized columns.

    Complexity
    ----------
    Time: O(n*m) where n=rows, m=columns
    Space: O(n*m) as data is loaded in memory
    """
    # Type validation
    if not isinstance(input_file, (str, Path)):
        raise TypeError(
            f"input_file must be str or Path, got {type(input_file).__name__}"
        )
    if not isinstance(output_file, (str, Path)):
        raise TypeError(
            f"output_file must be str or Path, got {type(output_file).__name__}"
        )
    if not isinstance(sheet_name, str):
        raise TypeError(f"sheet_name must be str, got {type(sheet_name).__name__}")
    if columns is not None and not isinstance(columns, list):
        raise TypeError(f"columns must be list or None, got {type(columns).__name__}")
    if max_rows is not None and not isinstance(max_rows, int):
        raise TypeError(f"max_rows must be int or None, got {type(max_rows).__name__}")
    if not isinstance(auto_format, bool):
        raise TypeError(f"auto_format must be bool, got {type(auto_format).__name__}")

    # Value validation
    if not Path(input_file).exists():
        raise FileNotFoundError(f"Input file not found: {input_file}")
    if not sheet_name:
        raise ValueError("sheet_name cannot be empty")
    if max_rows is not None and max_rows <= 0:
        raise ValueError(f"max_rows must be positive, got {max_rows}")

    EXCEL_MAX_ROWS = 1048576

    # Read Parquet file
    table = pq.read_table(input_file, columns=columns)

    # Validate columns if specified
    if columns is not None:
        for col in columns:
            if col not in table.schema.names:
                raise ValueError(
                    f"Column '{col}' not found in Parquet file. "
                    f"Available columns: {table.schema.names}"
                )

    # Convert to list of dicts
    data = table.to_pylist()

    # Apply row limit
    if max_rows is not None:
        data = data[:max_rows]

    # Check Excel row limit
    if len(data) >= EXCEL_MAX_ROWS:
        raise ValueError(
            f"Data has {len(data)} rows which exceeds Excel limit of {EXCEL_MAX_ROWS}. "
            f"Use max_rows parameter to limit output."
        )

    # Create workbook
    wb = Workbook()
    ws = wb.active
    if ws is None:
        ws = wb.create_sheet(sheet_name)
    else:
        ws.title = sheet_name

    if len(data) == 0:
        wb.save(output_file)
        return 0

    # Get column names
    column_names = list(data[0].keys())

    # Write header
    ws.append(column_names)

    # Write data rows
    for row_data in data:
        row_values = [row_data.get(col) for col in column_names]
        ws.append(row_values)

    # Apply formatting
    if auto_format:
        # Bold header
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(
                start_color="CCCCCC", end_color="CCCCCC", fill_type="solid"
            )

        # Freeze header row
        ws.freeze_panes = "A2"

        # Auto-size columns
        for col_idx, col_name in enumerate(column_names, start=1):
            max_length = len(str(col_name))

            for row_idx in range(2, min(len(data) + 2, 1000)):  # Sample first 1000 rows
                cell_value = ws.cell(row=row_idx, column=col_idx).value
                if cell_value is not None:
                    max_length = max(max_length, len(str(cell_value)))

            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[
                openpyxl.utils.get_column_letter(col_idx)
            ].width = adjusted_width

    # Save workbook
    wb.save(output_file)

    return len(data)


__all__ = ["parquet_to_excel"]

"""
Excel to Parquet converter.

Direct conversion from Excel to Parquet format with type inference.
"""

from pathlib import Path
from typing import Any

import openpyxl
import pyarrow as pa
import pyarrow.parquet as pq


def excel_to_parquet(
    input_file: str | Path,
    output_file: str | Path,
    *,
    sheet_name: str | None = None,
    skip_rows: int = 0,
    max_rows: int | None = None,
    columns: list[str] | None = None,
    type_inference: bool = True,
    compression: str = 'snappy',
) -> int:
    """
    Convert Excel sheet to Parquet format.
    
    Reads Excel data and converts to Parquet with automatic type inference.
    Provides efficient columnar storage for Excel data analysis.
    
    Parameters
    ----------
    input_file : str | Path
        Path to Excel file (.xlsx).
    output_file : str | Path
        Path to output Parquet file.
    sheet_name : str | None, optional
        Sheet to convert (by default None for active sheet).
    skip_rows : int, optional
        Number of rows to skip from top (by default 0).
    max_rows : int | None, optional
        Maximum rows to read (by default None for all).
    columns : list[str] | None, optional
        Specific columns to include (by default None for all).
    type_inference : bool, optional
        Automatically infer data types (by default True).
    compression : str, optional
        Compression codec: 'snappy', 'gzip', 'brotli', 'none' (by default 'snappy').
    
    Returns
    -------
    int
        Number of rows written to Parquet file.
    
    Raises
    ------
    TypeError
        If parameters are of wrong type.
    FileNotFoundError
        If input file doesn't exist.
    ValueError
        If parameters have invalid values.
    
    Examples
    --------
    >>> # Basic conversion
    >>> rows = excel_to_parquet('data.xlsx', 'output.parquet')
    >>> rows
    1000
    
    >>> # Convert specific sheet
    >>> rows = excel_to_parquet(
    ...     'workbook.xlsx',
    ...     'sales.parquet',
    ...     sheet_name='Sales'
    ... )
    
    >>> # Skip header rows and limit
    >>> rows = excel_to_parquet(
    ...     'data.xlsx',
    ...     'output.parquet',
    ...     skip_rows=2,
    ...     max_rows=500
    ... )
    
    >>> # Select columns with compression
    >>> rows = excel_to_parquet(
    ...     'data.xlsx',
    ...     'output.parquet',
    ...     columns=['ID', 'Name', 'Value'],
    ...     compression='gzip'
    ... )
    
    Notes
    -----
    Type inference converts Excel data types to appropriate Parquet types.
    Formulas are evaluated to their calculated values.
    
    Complexity
    ----------
    Time: O(n*m) where n=rows, m=columns
    Space: O(n*m) for data buffer
    """
    # Type validation
    if not isinstance(input_file, (str, Path)):
        raise TypeError(
            f"input_file must be str or Path, got {type(input_file).__name__}"
        )
    if not isinstance(output_file, (str, Path)):
        raise TypeError(
            f"output_file must be str or Path, got {type(output_file).__name__}"
        )
    if sheet_name is not None and not isinstance(sheet_name, str):
        raise TypeError(
            f"sheet_name must be str or None, got {type(sheet_name).__name__}"
        )
    if not isinstance(skip_rows, int):
        raise TypeError(f"skip_rows must be int, got {type(skip_rows).__name__}")
    if max_rows is not None and not isinstance(max_rows, int):
        raise TypeError(
            f"max_rows must be int or None, got {type(max_rows).__name__}"
        )
    if columns is not None and not isinstance(columns, list):
        raise TypeError(
            f"columns must be list or None, got {type(columns).__name__}"
        )
    if not isinstance(type_inference, bool):
        raise TypeError(
            f"type_inference must be bool, got {type(type_inference).__name__}"
        )
    if not isinstance(compression, str):
        raise TypeError(f"compression must be str, got {type(compression).__name__}")
    
    # Value validation
    input_path = Path(input_file)
    if not input_path.exists():
        raise FileNotFoundError(f"Input file not found: {input_file}")
    
    if skip_rows < 0:
        raise ValueError(f"skip_rows must be non-negative, got {skip_rows}")
    
    if max_rows is not None and max_rows <= 0:
        raise ValueError(f"max_rows must be positive, got {max_rows}")
    
    valid_compression = {'snappy', 'gzip', 'brotli', 'zstd', 'lz4', 'none'}
    if compression not in valid_compression:
        raise ValueError(
            f"compression must be one of {valid_compression}, got '{compression}'"
        )
    
    # Load workbook
    wb = openpyxl.load_workbook(input_path, read_only=True, data_only=True)
    
    try:
        # Select sheet
        if sheet_name is None:
            ws = wb.active
        else:
            if sheet_name not in wb.sheetnames:
                raise ValueError(f"Sheet '{sheet_name}' not found in workbook")
            ws = wb[sheet_name]
        
        # Read data
        rows_iter = ws.iter_rows(values_only=True)
        
        # Skip initial rows
        for _ in range(skip_rows):
            next(rows_iter, None)
        
        # Get header row
        header = next(rows_iter, None)
        if header is None:
            return 0
        
        # Clean header (convert None to empty string)
        header = [str(col) if col is not None else f'Column{i}' 
                 for i, col in enumerate(header)]
        
        # Filter columns if specified
        if columns is not None:
            col_indices = []
            for col in columns:
                if col not in header:
                    raise ValueError(f"Column '{col}' not found in sheet")
                col_indices.append(header.index(col))
            header = columns
        else:
            col_indices = list(range(len(header)))
        
        # Read data rows
        data_rows = []
        row_count = 0
        
        for row in rows_iter:
            if max_rows is not None and row_count >= max_rows:
                break
            
            # Filter columns
            filtered_row = [row[i] if i < len(row) else None for i in col_indices]
            data_rows.append(filtered_row)
            row_count += 1
        
        if row_count == 0:
            # Create empty table with schema
            fields = [pa.field(col, pa.string()) for col in header]
            schema = pa.schema(fields)
            table = pa.Table.from_pydict({col: [] for col in header}, schema=schema)
        else:
            # Convert to dictionary format
            data_dict = {col: [row[i] for row in data_rows] 
                        for i, col in enumerate(header)}
            
            # Create PyArrow table
            if type_inference:
                table = pa.Table.from_pydict(data_dict)
            else:
                # Use string type for all columns
                fields = [pa.field(col, pa.string()) for col in header]
                schema = pa.schema(fields)
                # Convert all values to strings
                string_dict = {col: [str(v) if v is not None else None for v in vals]
                              for col, vals in data_dict.items()}
                table = pa.Table.from_pydict(string_dict, schema=schema)
        
        # Write Parquet file
        output_path = Path(output_file)
        output_path.parent.mkdir(parents=True, exist_ok=True)
        
        pq.write_table(table, output_path, compression=compression)
        
        return row_count
    
    finally:
        wb.close()


__all__ = ['excel_to_parquet']

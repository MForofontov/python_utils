"""
Excel to CSV batch converter.

Converts all sheets from an Excel workbook to separate CSV files.
"""

from pathlib import Path
from typing import Any

import openpyxl


def excel_to_csv_batch(
    input_file: str | Path,
    output_dir: str | Path,
    *,
    sheet_names: list[str] | None = None,
    skip_empty: bool = True,
    encoding: str = 'utf-8',
) -> dict[str, int]:
    """
    Convert Excel workbook sheets to separate CSV files.
    
    Each sheet is exported to a CSV file named after the sheet. Useful for
    batch processing Excel data or migrating to CSV-based workflows.
    
    Parameters
    ----------
    input_file : str | Path
        Path to Excel file (.xlsx) to convert.
    output_dir : str | Path
        Directory where CSV files will be created.
    sheet_names : list[str] | None, optional
        Specific sheets to convert (by default None for all sheets).
    skip_empty : bool, optional
        Skip sheets with no data (by default True).
    encoding : str, optional
        Character encoding for CSV files (by default 'utf-8').
    
    Returns
    -------
    dict[str, int]
        Mapping of sheet names to row counts exported.
    
    Raises
    ------
    TypeError
        If parameters are of wrong type.
    FileNotFoundError
        If input file doesn't exist.
    ValueError
        If parameters have invalid values.
    
    Examples
    --------
    >>> # Convert all sheets
    >>> counts = excel_to_csv_batch('data.xlsx', 'csv_output/')
    >>> counts
    {'Sheet1': 100, 'Sheet2': 50, 'Sheet3': 75}
    
    >>> # Convert specific sheets only
    >>> counts = excel_to_csv_batch(
    ...     'data.xlsx',
    ...     'csv_output/',
    ...     sheet_names=['Sales', 'Inventory']
    ... )
    
    >>> # Include empty sheets
    >>> counts = excel_to_csv_batch(
    ...     'data.xlsx',
    ...     'csv_output/',
    ...     skip_empty=False
    ... )
    
    Notes
    -----
    Output CSV files are named: {sheet_name}.csv
    Special characters in sheet names are replaced with underscores.
    
    Complexity
    ----------
    Time: O(n*m) where n=sheets, m=rows per sheet
    Space: O(m) for largest sheet
    """
    # Type validation
    if not isinstance(input_file, (str, Path)):
        raise TypeError(
            f"input_file must be str or Path, got {type(input_file).__name__}"
        )
    if not isinstance(output_dir, (str, Path)):
        raise TypeError(
            f"output_dir must be str or Path, got {type(output_dir).__name__}"
        )
    if sheet_names is not None and not isinstance(sheet_names, list):
        raise TypeError(
            f"sheet_names must be list or None, got {type(sheet_names).__name__}"
        )
    if not isinstance(skip_empty, bool):
        raise TypeError(f"skip_empty must be bool, got {type(skip_empty).__name__}")
    if not isinstance(encoding, str):
        raise TypeError(f"encoding must be str, got {type(encoding).__name__}")
    
    # Value validation
    input_path = Path(input_file)
    if not input_path.exists():
        raise FileNotFoundError(f"Input file not found: {input_file}")
    
    output_path = Path(output_dir)
    output_path.mkdir(parents=True, exist_ok=True)
    
    # Load workbook
    wb = openpyxl.load_workbook(input_path, read_only=True, data_only=True)
    
    # Determine sheets to process
    if sheet_names is None:
        sheets_to_process = wb.sheetnames
    else:
        sheets_to_process = sheet_names
        # Validate sheet names exist
        for name in sheets_to_process:
            if name not in wb.sheetnames:
                wb.close()
                raise ValueError(f"Sheet '{name}' not found in workbook")
    
    results: dict[str, int] = {}
    
    try:
        for sheet_name in sheets_to_process:
            ws = wb[sheet_name]
            
            # Check if sheet is empty (has no data rows when iterating)
            if skip_empty:
                # Quick check - see if sheet has any actual data
                has_data = False
                for row in ws.iter_rows(values_only=True):
                    if not all(cell is None for cell in row):
                        has_data = True
                        break
                if not has_data:
                    continue  # Skip this sheet entirely
            
            # Create safe filename
            safe_name = "".join(c if c.isalnum() or c in ('-', '_') else '_' 
                               for c in sheet_name)
            csv_file = output_path / f"{safe_name}.csv"
            
            # Write CSV
            import csv
            row_count = 0
            with open(csv_file, 'w', newline='', encoding=encoding) as f:
                writer = csv.writer(f)
                
                for row in ws.iter_rows(values_only=True):
                    # Skip completely empty rows
                    if skip_empty and all(cell is None for cell in row):
                        continue
                    
                    # Convert None to empty string for CSV
                    cleaned_row = ['' if cell is None else str(cell) for cell in row]
                    writer.writerow(cleaned_row)
                    row_count += 1
            
            results[sheet_name] = row_count
    
    finally:
        wb.close()
    
    return results


__all__ = ['excel_to_csv_batch']

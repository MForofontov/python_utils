"""Excel structure validation."""

from pathlib import Path
from typing import Any

import openpyxl


def validate_excel_structure(
    file_path: str | Path,
    *,
    sheet_name: str | None = None,
    required_columns: list[str] | None = None,
    min_rows: int = 0,
    max_rows: int | None = None,
    strict_columns: bool = False,
) -> dict[str, Any]:
    """
    Validate Excel sheet structure against requirements.

    Checks Excel sheet for required columns, row counts, and data integrity.
    Returns detailed validation results with errors and warnings.

    Parameters
    ----------
    file_path : str | Path
        Path to Excel file to validate.
    sheet_name : str | None, optional
        Sheet to validate (by default None for active sheet).
    required_columns : list[str] | None, optional
        Required column names (by default None).
    min_rows : int, optional
        Minimum required data rows (by default 0).
    max_rows : int | None, optional
        Maximum allowed data rows (by default None for unlimited).
    strict_columns : bool, optional
        Require exact column match with no extras (by default False).

    Returns
    -------
    dict[str, Any]
        Validation results with keys:
        - 'valid': bool - Overall validation status
        - 'row_count': int - Number of data rows
        - 'columns': list[str] - Found column names
        - 'errors': list[str] - Validation errors
        - 'warnings': list[str] - Validation warnings

    Raises
    ------
    TypeError
        If parameters are of wrong type.
    FileNotFoundError
        If file doesn't exist.
    ValueError
        If parameters have invalid values.

    Examples
    --------
    >>> # Basic validation
    >>> result = validate_excel_structure('data.xlsx')
    >>> result['valid']
    True

    >>> # Check required columns
    >>> result = validate_excel_structure(
    ...     'data.xlsx',
    ...     required_columns=['ID', 'Name', 'Email']
    ... )
    >>> result['errors']
    []

    >>> # Validate row count range
    >>> result = validate_excel_structure(
    ...     'data.xlsx',
    ...     min_rows=10,
    ...     max_rows=1000
    ... )

    >>> # Strict column matching
    >>> result = validate_excel_structure(
    ...     'data.xlsx',
    ...     required_columns=['ID', 'Name'],
    ...     strict_columns=True
    ... )

    Notes
    -----
    Validation checks:
    - Column presence (if required_columns specified)
    - Row count within limits
    - Empty/duplicate columns
    - Formula errors in cells

    Complexity
    ----------
    Time: O(n*m) where n=rows, m=columns
    Space: O(m) for column storage
    """
    # Type validation
    if not isinstance(file_path, (str, Path)):
        raise TypeError(
            f"file_path must be str or Path, got {type(file_path).__name__}"
        )
    if sheet_name is not None and not isinstance(sheet_name, str):
        raise TypeError(
            f"sheet_name must be str or None, got {type(sheet_name).__name__}"
        )
    if required_columns is not None and not isinstance(required_columns, list):
        raise TypeError(
            f"required_columns must be list or None, got {type(required_columns).__name__}"
        )
    if not isinstance(min_rows, int):
        raise TypeError(f"min_rows must be int, got {type(min_rows).__name__}")
    if max_rows is not None and not isinstance(max_rows, int):
        raise TypeError(f"max_rows must be int or None, got {type(max_rows).__name__}")
    if not isinstance(strict_columns, bool):
        raise TypeError(
            f"strict_columns must be bool, got {type(strict_columns).__name__}"
        )

    # Value validation
    file_path = Path(file_path)
    if not file_path.exists():
        raise FileNotFoundError(f"Excel file not found: {file_path}")

    if min_rows < 0:
        raise ValueError(f"min_rows must be non-negative, got {min_rows}")

    if max_rows is not None and max_rows < min_rows:
        raise ValueError(f"max_rows ({max_rows}) must be >= min_rows ({min_rows})")

    # Load workbook
    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=False)

    result: dict[str, Any] = {
        "valid": True,
        "row_count": 0,
        "columns": [],
        "errors": [],
        "warnings": [],
    }

    try:
        # Select sheet
        if sheet_name is None:
            ws = wb.active
        else:
            if sheet_name not in wb.sheetnames:
                result["errors"].append(f"Sheet '{sheet_name}' not found")
                result["valid"] = False
                return result
            ws = wb[sheet_name]

        # Get header row
        rows_iter = ws.iter_rows(values_only=True)
        header = next(rows_iter, None)

        if header is None:
            result["errors"].append("Sheet is empty - no header row found")
            result["valid"] = False
            return result

        # Clean and validate header
        columns = []
        for i, col in enumerate(header):
            if col is None:
                result["warnings"].append(f"Empty column header at position {i}")
                columns.append(f"Column{i}")
            else:
                col_str = str(col).strip()
                if not col_str:
                    result["warnings"].append(f"Empty column header at position {i}")
                    columns.append(f"Column{i}")
                else:
                    columns.append(col_str)

        result["columns"] = columns

        # Check for duplicate columns
        seen_cols: set[str] = set()
        for col in columns:
            if col in seen_cols:
                result["warnings"].append(f"Duplicate column name: '{col}'")
            seen_cols.add(col)

        # Validate required columns
        if required_columns is not None:
            missing_cols = set(required_columns) - set(columns)
            if missing_cols:
                result["errors"].append(
                    f"Missing required columns: {sorted(missing_cols)}"
                )
                result["valid"] = False

        # Check strict column matching
        if strict_columns and required_columns is not None:
            extra_cols = set(columns) - set(required_columns)
            if extra_cols:
                result["errors"].append(
                    f"Extra columns not allowed (strict mode): {sorted(extra_cols)}"
                )
                result["valid"] = False

        # Count data rows
        row_count = 0
        for _ in rows_iter:
            row_count += 1

        result["row_count"] = row_count

        # Validate row count
        if row_count < min_rows:
            result["errors"].append(
                f"Insufficient rows: found {row_count}, minimum required {min_rows}"
            )
            result["valid"] = False

        if max_rows is not None and row_count > max_rows:
            result["errors"].append(
                f"Too many rows: found {row_count}, maximum allowed {max_rows}"
            )
            result["valid"] = False

        # Check for formula errors (reload with data_only=False)
        wb_formulas = openpyxl.load_workbook(file_path, read_only=True, data_only=False)
        ws_formulas = wb_formulas[ws.title]

        error_count = 0
        for row in ws_formulas.iter_rows(min_row=2):  # Skip header
            for cell in row:
                if isinstance(cell.value, str) and cell.value.startswith("#"):
                    error_count += 1
                    if error_count <= 3:  # Limit error messages
                        result["warnings"].append(
                            f"Formula error in cell {cell.coordinate}: {cell.value}"
                        )

        if error_count > 3:
            result["warnings"].append(f"... and {error_count - 3} more formula errors")

        wb_formulas.close()

    finally:
        wb.close()

    return result


__all__ = ["validate_excel_structure"]

"""
Unit tests for parquet_to_excel function.
"""

import tempfile
from pathlib import Path

import pytest

try:
    import openpyxl
    import pyarrow as pa
    import pyarrow.parquet as pq
    from python_utils.serialization_functions.format_converters.parquet_to_excel import parquet_to_excel
    DEPENDENCIES_AVAILABLE = True
except ImportError:
    DEPENDENCIES_AVAILABLE = False
    openpyxl = None  # type: ignore
    pa = None  # type: ignore
    pq = None  # type: ignore
    parquet_to_excel = None  # type: ignore

pytestmark = pytest.mark.skipif(not DEPENDENCIES_AVAILABLE, reason="openpyxl/pyarrow not installed")
pytestmark = [pytestmark, pytest.mark.unit, pytest.mark.serialization]


def test_parquet_to_excel_basic() -> None:
    """
    Test basic Parquet to Excel conversion.
    """
    with tempfile.TemporaryDirectory() as tmpdir:
        parquet_file = Path(tmpdir) / "data.parquet"
        excel_file = Path(tmpdir) / "data.xlsx"

        data = [
            {"id": 1, "name": "Alice", "age": 30},
            {"id": 2, "name": "Bob", "age": 25},
        ]

        table = pa.Table.from_pylist(data)
        pq.write_table(table, parquet_file)

        rows = parquet_to_excel(parquet_file, excel_file)

        assert rows == 2
        assert excel_file.exists()

        # Verify Excel data
        wb = openpyxl.load_workbook(excel_file)
        ws = wb.active
        data_rows = list(ws.iter_rows(values_only=True))
        assert len(data_rows) == 3  # header + 2 rows
        assert data_rows[0] == ("id", "name", "age")


def test_parquet_to_excel_select_columns() -> None:
    """
    Test exporting specific columns only.
    """
    with tempfile.TemporaryDirectory() as tmpdir:
        parquet_file = Path(tmpdir) / "data.parquet"
        excel_file = Path(tmpdir) / "data.xlsx"

        data = [{"id": 1, "name": "Alice", "age": 30, "email": "alice@example.com"}]

        table = pa.Table.from_pylist(data)
        pq.write_table(table, parquet_file)

        rows = parquet_to_excel(parquet_file, excel_file, columns=["id", "name"])

        assert rows == 1

        wb = openpyxl.load_workbook(excel_file)
        ws = wb.active
        header = next(ws.iter_rows(values_only=True))
        assert header == ("id", "name")


def test_parquet_to_excel_max_rows() -> None:
    """
    Test limiting rows in export.
    """
    with tempfile.TemporaryDirectory() as tmpdir:
        parquet_file = Path(tmpdir) / "data.parquet"
        excel_file = Path(tmpdir) / "data.xlsx"

        data = [{"id": i} for i in range(100)]

        table = pa.Table.from_pylist(data)
        pq.write_table(table, parquet_file)

        rows = parquet_to_excel(parquet_file, excel_file, max_rows=10)

        assert rows == 10


def test_parquet_to_excel_custom_sheet_name() -> None:
    """
    Test custom sheet name.
    """
    with tempfile.TemporaryDirectory() as tmpdir:
        parquet_file = Path(tmpdir) / "data.parquet"
        excel_file = Path(tmpdir) / "data.xlsx"

        data = [{"id": 1}]
        table = pa.Table.from_pylist(data)
        pq.write_table(table, parquet_file)

        parquet_to_excel(parquet_file, excel_file, sheet_name="MyData")

        wb = openpyxl.load_workbook(excel_file)
        assert wb.active.title == "MyData"


def test_parquet_to_excel_with_formatting() -> None:
    """
    Test auto-formatting is applied.
    """
    with tempfile.TemporaryDirectory() as tmpdir:
        parquet_file = Path(tmpdir) / "data.parquet"
        excel_file = Path(tmpdir) / "data.xlsx"

        data = [{"id": 1, "name": "Alice"}]
        table = pa.Table.from_pylist(data)
        pq.write_table(table, parquet_file)

        parquet_to_excel(parquet_file, excel_file, auto_format=True)

        wb = openpyxl.load_workbook(excel_file)
        ws = wb.active

        # Check formatting
        assert ws.freeze_panes == "A2"
        assert ws.cell(1, 1).font.bold is True


def test_parquet_to_excel_no_formatting() -> None:
    """
    Test without auto-formatting.
    """
    with tempfile.TemporaryDirectory() as tmpdir:
        parquet_file = Path(tmpdir) / "data.parquet"
        excel_file = Path(tmpdir) / "data.xlsx"

        data = [{"id": 1}]
        table = pa.Table.from_pylist(data)
        pq.write_table(table, parquet_file)

        parquet_to_excel(parquet_file, excel_file, auto_format=False)

        wb = openpyxl.load_workbook(excel_file)
        ws = wb.active

        assert ws.freeze_panes is None


def test_parquet_to_excel_empty_data() -> None:
    """
    Test handling empty Parquet file.
    """
    with tempfile.TemporaryDirectory() as tmpdir:
        parquet_file = Path(tmpdir) / "data.parquet"
        excel_file = Path(tmpdir) / "data.xlsx"

        data: list[dict] = []
        schema = pa.schema([("id", pa.int64())])
        table = pa.Table.from_pylist(data, schema=schema)
        pq.write_table(table, parquet_file)

        rows = parquet_to_excel(parquet_file, excel_file)

        assert rows == 0


def test_parquet_to_excel_invalid_type_input() -> None:
    """
    Test TypeError for invalid input_file type.
    """
    with pytest.raises(TypeError, match="input_file must be str or Path"):
        parquet_to_excel(123, "output.xlsx")  # type: ignore


def test_parquet_to_excel_invalid_max_rows() -> None:
    """
    Test ValueError for invalid max_rows value.
    """
    with tempfile.TemporaryDirectory() as tmpdir:
        parquet_file = Path(tmpdir) / "data.parquet"
        excel_file = Path(tmpdir) / "data.xlsx"

        data = [{"id": 1}]
        table = pa.Table.from_pylist(data)
        pq.write_table(table, parquet_file)

        with pytest.raises(ValueError, match="max_rows must be positive"):
            parquet_to_excel(parquet_file, excel_file, max_rows=0)


def test_parquet_to_excel_file_not_found() -> None:
    """
    Test FileNotFoundError for non-existent file.
    """
    with tempfile.TemporaryDirectory() as tmpdir:
        excel_file = Path(tmpdir) / "output.xlsx"

        with pytest.raises(FileNotFoundError, match="Input file not found"):
            parquet_to_excel("/nonexistent/file.parquet", excel_file)


def test_parquet_to_excel_invalid_column() -> None:
    """
    Test error for non-existent column.
    """
    with tempfile.TemporaryDirectory() as tmpdir:
        parquet_file = Path(tmpdir) / "data.parquet"
        excel_file = Path(tmpdir) / "data.xlsx"

        data = [{"id": 1, "name": "Alice"}]
        table = pa.Table.from_pylist(data)
        pq.write_table(table, parquet_file)

        # PyArrow raises ArrowInvalid for non-existent columns
        with pytest.raises((ValueError, Exception), match="nonexistent|No match"):
            parquet_to_excel(parquet_file, excel_file, columns=["nonexistent"])

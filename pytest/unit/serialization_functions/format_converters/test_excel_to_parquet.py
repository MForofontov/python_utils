"""
Unit tests for excel_to_parquet function.
"""

import tempfile
from pathlib import Path

try:
    import openpyxl
    import pyarrow
    import pyarrow.parquet as pq
    DEPENDENCIES_AVAILABLE = True
except ImportError:
    DEPENDENCIES_AVAILABLE = False
    openpyxl = None  # type: ignore
    pyarrow = None  # type: ignore
    pq = None  # type: ignore

import pytest
from serialization_functions.format_converters.excel_to_parquet import excel_to_parquet

pytestmark = pytest.mark.skipif(not DEPENDENCIES_AVAILABLE, reason="openpyxl/pyarrow not installed")
pytestmark = [pytestmark, pytest.mark.unit, pytest.mark.serialization]


def test_excel_to_parquet_basic() -> None:
    """
    Test basic Excel to Parquet conversion.
    """
    with tempfile.TemporaryDirectory() as tmpdir:
        excel_file = Path(tmpdir) / "data.xlsx"
        parquet_file = Path(tmpdir) / "output.parquet"

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["ID", "Name", "Age"])
        ws.append([1, "Alice", 30])
        ws.append([2, "Bob", 25])

        wb.save(excel_file)
        wb.close()

        rows = excel_to_parquet(excel_file, parquet_file)

        assert rows == 2
        assert parquet_file.exists()

        # Verify Parquet content
        table = pq.read_table(parquet_file)
        assert len(table) == 2


def test_excel_to_parquet_specific_sheet() -> None:
    """
    Test converting specific sheet.
    """
    with tempfile.TemporaryDirectory() as tmpdir:
        excel_file = Path(tmpdir) / "data.xlsx"
        parquet_file = Path(tmpdir) / "output.parquet"

        wb = openpyxl.Workbook()
        ws1 = wb.active
        ws1.title = "Sheet1"
        ws1.append(["Data"])
        ws1.append([1])

        ws2 = wb.create_sheet("Sheet2")
        ws2.append(["Other"])
        ws2.append([99])

        wb.save(excel_file)
        wb.close()

        excel_to_parquet(excel_file, parquet_file, sheet_name="Sheet2")

        table = pq.read_table(parquet_file)
        data = table.to_pylist()
        assert data[0]["Other"] == 99


def test_excel_to_parquet_skip_rows() -> None:
    """
    Test skipping initial rows.
    """
    with tempfile.TemporaryDirectory() as tmpdir:
        excel_file = Path(tmpdir) / "data.xlsx"
        parquet_file = Path(tmpdir) / "output.parquet"

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Title"])
        ws.append(["Subtitle"])
        ws.append(["ID", "Name"])  # Real header
        ws.append([1, "Alice"])

        wb.save(excel_file)
        wb.close()

        rows = excel_to_parquet(excel_file, parquet_file, skip_rows=2)

        assert rows == 1
        table = pq.read_table(parquet_file)
        assert "ID" in table.column_names


def test_excel_to_parquet_max_rows() -> None:
    """
    Test limiting maximum rows.
    """
    with tempfile.TemporaryDirectory() as tmpdir:
        excel_file = Path(tmpdir) / "data.xlsx"
        parquet_file = Path(tmpdir) / "output.parquet"

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["ID"])
        for i in range(100):
            ws.append([i])

        wb.save(excel_file)
        wb.close()

        rows = excel_to_parquet(excel_file, parquet_file, max_rows=10)

        assert rows == 10


def test_excel_to_parquet_select_columns() -> None:
    """
    Test selecting specific columns.
    """
    with tempfile.TemporaryDirectory() as tmpdir:
        excel_file = Path(tmpdir) / "data.xlsx"
        parquet_file = Path(tmpdir) / "output.parquet"

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["ID", "Name", "Email"])
        ws.append([1, "Alice", "alice@example.com"])

        wb.save(excel_file)
        wb.close()

        excel_to_parquet(excel_file, parquet_file, columns=["ID", "Name"])

        table = pq.read_table(parquet_file)
        assert "ID" in table.column_names
        assert "Name" in table.column_names
        assert "Email" not in table.column_names


def test_excel_to_parquet_custom_compression() -> None:
    """
    Test custom compression codec.
    """
    with tempfile.TemporaryDirectory() as tmpdir:
        excel_file = Path(tmpdir) / "data.xlsx"
        parquet_file = Path(tmpdir) / "output.parquet"

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["ID"])
        ws.append([1])

        wb.save(excel_file)
        wb.close()

        excel_to_parquet(excel_file, parquet_file, compression="gzip")

        assert parquet_file.exists()


def test_excel_to_parquet_no_type_inference() -> None:
    """
    Test without type inference (all strings).
    """
    with tempfile.TemporaryDirectory() as tmpdir:
        excel_file = Path(tmpdir) / "data.xlsx"
        parquet_file = Path(tmpdir) / "output.parquet"

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["ID", "Value"])
        ws.append([1, 100])

        wb.save(excel_file)
        wb.close()

        excel_to_parquet(excel_file, parquet_file, type_inference=False)

        table = pq.read_table(parquet_file)
        # All columns should be string type
        for field in table.schema:
            assert str(field.type) == "string"


def test_excel_to_parquet_empty_sheet() -> None:
    """
    Test handling empty sheet.
    """
    with tempfile.TemporaryDirectory() as tmpdir:
        excel_file = Path(tmpdir) / "data.xlsx"
        parquet_file = Path(tmpdir) / "output.parquet"

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["ID"])
        # No data rows

        wb.save(excel_file)
        wb.close()

        rows = excel_to_parquet(excel_file, parquet_file)

        assert rows == 0


def test_excel_to_parquet_invalid_type_input() -> None:
    """
    Test TypeError for invalid input_file type.
    """
    with pytest.raises(TypeError, match="input_file must be str or Path"):
        excel_to_parquet(123, "output.parquet")  # type: ignore


def test_excel_to_parquet_invalid_skip_rows() -> None:
    """
    Test ValueError for negative skip_rows.
    """
    with tempfile.TemporaryDirectory() as tmpdir:
        excel_file = Path(tmpdir) / "data.xlsx"
        parquet_file = Path(tmpdir) / "output.parquet"

        wb = openpyxl.Workbook()
        wb.save(excel_file)
        wb.close()

        with pytest.raises(ValueError, match="skip_rows must be non-negative"):
            excel_to_parquet(excel_file, parquet_file, skip_rows=-1)


def test_excel_to_parquet_file_not_found() -> None:
    """
    Test FileNotFoundError for non-existent file.
    """
    with tempfile.TemporaryDirectory() as tmpdir:
        parquet_file = Path(tmpdir) / "output.parquet"

        with pytest.raises(FileNotFoundError, match="Input file not found"):
            excel_to_parquet("/nonexistent/file.xlsx", parquet_file)


def test_excel_to_parquet_invalid_compression() -> None:
    """
    Test ValueError for invalid compression codec.
    """
    with tempfile.TemporaryDirectory() as tmpdir:
        excel_file = Path(tmpdir) / "data.xlsx"
        parquet_file = Path(tmpdir) / "output.parquet"

        wb = openpyxl.Workbook()
        wb.save(excel_file)
        wb.close()

        with pytest.raises(ValueError, match="compression must be one of"):
            excel_to_parquet(excel_file, parquet_file, compression="invalid")


def test_excel_to_parquet_invalid_sheet_name() -> None:
    """
    Test ValueError for non-existent sheet.
    """
    with tempfile.TemporaryDirectory() as tmpdir:
        excel_file = Path(tmpdir) / "data.xlsx"
        parquet_file = Path(tmpdir) / "output.parquet"

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        wb.save(excel_file)
        wb.close()

        with pytest.raises(ValueError, match="not found in workbook"):
            excel_to_parquet(excel_file, parquet_file, sheet_name="NonExistent")


def test_excel_to_parquet_invalid_column() -> None:
    """
    Test ValueError for non-existent column.
    """
    with tempfile.TemporaryDirectory() as tmpdir:
        excel_file = Path(tmpdir) / "data.xlsx"
        parquet_file = Path(tmpdir) / "output.parquet"

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["ID", "Name"])
        ws.append([1, "Alice"])

        wb.save(excel_file)
        wb.close()

        with pytest.raises(ValueError, match="not found in sheet"):
            excel_to_parquet(excel_file, parquet_file, columns=["NonExistent"])

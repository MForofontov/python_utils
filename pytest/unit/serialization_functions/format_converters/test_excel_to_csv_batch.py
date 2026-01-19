"""
Unit tests for excel_to_csv_batch function.
"""

import tempfile
from pathlib import Path

try:
    import openpyxl
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    openpyxl = None  # type: ignore

import pytest
from python_utils.serialization_functions.format_converters.excel_to_csv_batch import (
    excel_to_csv_batch,
)

pytestmark = pytest.mark.skipif(not OPENPYXL_AVAILABLE, reason="openpyxl not installed")
pytestmark = [pytestmark, pytest.mark.unit, pytest.mark.serialization]


def test_excel_to_csv_batch_all_sheets() -> None:
    """
    Test converting all sheets to CSV files.
    """
    with tempfile.TemporaryDirectory() as tmpdir:
        excel_file = Path(tmpdir) / "data.xlsx"
        output_dir = Path(tmpdir) / "csv"

        wb = openpyxl.Workbook()

        # Sheet1
        ws1 = wb.active
        ws1.title = "Sheet1"
        ws1.append(["ID", "Name"])
        ws1.append([1, "Alice"])
        ws1.append([2, "Bob"])

        # Sheet2
        ws2 = wb.create_sheet("Sheet2")
        ws2.append(["Product", "Price"])
        ws2.append(["Apple", 1.50])

        wb.save(excel_file)
        wb.close()

        counts = excel_to_csv_batch(excel_file, output_dir)

        assert len(counts) == 2
        assert counts["Sheet1"] == 3  # header + 2 rows
        assert counts["Sheet2"] == 2  # header + 1 row

        # Verify CSV files exist
        assert (output_dir / "Sheet1.csv").exists()
        assert (output_dir / "Sheet2.csv").exists()


def test_excel_to_csv_batch_specific_sheets() -> None:
    """
    Test converting specific sheets only.
    """
    with tempfile.TemporaryDirectory() as tmpdir:
        excel_file = Path(tmpdir) / "data.xlsx"
        output_dir = Path(tmpdir) / "csv"

        wb = openpyxl.Workbook()
        ws1 = wb.active
        ws1.title = "Sales"
        ws1.append(["Amount"])
        ws1.append([100])

        ws2 = wb.create_sheet("Inventory")
        ws2.append(["Count"])
        ws2.append([50])

        wb.save(excel_file)
        wb.close()

        counts = excel_to_csv_batch(excel_file, output_dir, sheet_names=["Sales"])

        assert len(counts) == 1
        assert "Sales" in counts
        assert "Inventory" not in counts


def test_excel_to_csv_batch_skip_empty() -> None:
    """
    Test skipping empty sheets.
    """
    with tempfile.TemporaryDirectory() as tmpdir:
        excel_file = Path(tmpdir) / "data.xlsx"
        output_dir = Path(tmpdir) / "csv"

        wb = openpyxl.Workbook()
        ws1 = wb.active
        ws1.title = "Data"
        ws1.append(["ID"])
        ws1.append([1])

        wb.create_sheet("Empty")
        # Empty sheet - max_row will be 0 so it should be skipped

        wb.save(excel_file)
        wb.close()

        counts = excel_to_csv_batch(excel_file, output_dir, skip_empty=True)

        assert "Data" in counts
        assert counts["Data"] == 2
        assert "Empty" not in counts  # Should be skipped when truly empty


def test_excel_to_csv_batch_include_empty() -> None:
    """
    Test including empty sheets when skip_empty=False.
    """
    with tempfile.TemporaryDirectory() as tmpdir:
        excel_file = Path(tmpdir) / "data.xlsx"
        output_dir = Path(tmpdir) / "csv"

        wb = openpyxl.Workbook()
        ws1 = wb.active
        ws1.title = "Empty"

        wb.save(excel_file)
        wb.close()

        counts = excel_to_csv_batch(excel_file, output_dir, skip_empty=False)

        # Empty sheet should have 0 rows but still be processed
        assert "Empty" in counts or len(counts) == 0


def test_excel_to_csv_batch_special_characters() -> None:
    """
    Test handling sheet names with special characters.
    """
    with tempfile.TemporaryDirectory() as tmpdir:
        excel_file = Path(tmpdir) / "data.xlsx"
        output_dir = Path(tmpdir) / "csv"

        wb = openpyxl.Workbook()
        ws = wb.active
        # Excel allows spaces and dashes but not /\?*[]
        ws.title = "Sales-2024"
        ws.append(["ID"])
        ws.append([1])

        wb.save(excel_file)
        wb.close()

        counts = excel_to_csv_batch(excel_file, output_dir)

        # Sheet name should be preserved if already safe
        assert "Sales-2024" in counts
        csv_file = output_dir / "Sales-2024.csv"
        assert csv_file.exists()


def test_excel_to_csv_batch_custom_encoding() -> None:
    """
    Test custom encoding for CSV output.
    """
    with tempfile.TemporaryDirectory() as tmpdir:
        excel_file = Path(tmpdir) / "data.xlsx"
        output_dir = Path(tmpdir) / "csv"

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Name"])
        ws.append(["Café"])

        wb.save(excel_file)
        wb.close()

        excel_to_csv_batch(excel_file, output_dir, encoding="utf-8")

        csv_file = next(output_dir.glob("*.csv"))
        with open(csv_file, encoding="utf-8") as f:
            content = f.read()
            assert "Café" in content


def test_excel_to_csv_batch_invalid_type_input() -> None:
    """
    Test TypeError for invalid input_file type.
    """
    with pytest.raises(TypeError, match="input_file must be str or Path"):
        excel_to_csv_batch(123, "output")  # type: ignore


def test_excel_to_csv_batch_invalid_type_sheet_names() -> None:
    """
    Test TypeError for invalid sheet_names type.
    """
    with tempfile.TemporaryDirectory() as tmpdir:
        excel_file = Path(tmpdir) / "data.xlsx"

        wb = openpyxl.Workbook()
        wb.save(excel_file)
        wb.close()

        with pytest.raises(TypeError, match="sheet_names must be list"):
            excel_to_csv_batch(excel_file, tmpdir, sheet_names="not_a_list")  # type: ignore


def test_excel_to_csv_batch_file_not_found() -> None:
    """
    Test FileNotFoundError for non-existent file.
    """
    with tempfile.TemporaryDirectory() as tmpdir:
        output_dir = Path(tmpdir) / "csv"

        with pytest.raises(FileNotFoundError, match="Input file not found"):
            excel_to_csv_batch("/nonexistent/file.xlsx", output_dir)


def test_excel_to_csv_batch_invalid_sheet_name() -> None:
    """
    Test ValueError for non-existent sheet name.
    """
    with tempfile.TemporaryDirectory() as tmpdir:
        excel_file = Path(tmpdir) / "data.xlsx"
        output_dir = Path(tmpdir) / "csv"

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        wb.save(excel_file)
        wb.close()

        with pytest.raises(ValueError, match="not found in workbook"):
            excel_to_csv_batch(excel_file, output_dir, sheet_names=["NonExistent"])

"""
Unit tests for merge_excel_sheets function.
"""

import tempfile
from pathlib import Path

try:
    import openpyxl
    from python_utils.serialization_functions.excel_operations.merge_excel_sheets import (
        merge_excel_sheets,
    )
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    openpyxl = None  # type: ignore
    merge_excel_sheets = None  # type: ignore

import pytest

pytestmark = pytest.mark.skipif(not OPENPYXL_AVAILABLE, reason="openpyxl not installed")
pytestmark = [pytestmark, pytest.mark.unit, pytest.mark.serialization]


def test_merge_excel_sheets_basic_merge() -> None:
    """
    Test basic merge of multiple sheets.
    """
    with tempfile.TemporaryDirectory() as tmpdir:
        input_file = Path(tmpdir) / "input.xlsx"
        output_file = Path(tmpdir) / "output.xlsx"

        # Create workbook with multiple sheets
        wb = openpyxl.Workbook()
        ws1 = wb.active
        ws1.title = "Sheet1"
        ws1.append(["id", "name"])
        ws1.append([1, "Alice"])
        ws1.append([2, "Bob"])

        ws2 = wb.create_sheet("Sheet2")
        ws2.append(["id", "name"])
        ws2.append([3, "Charlie"])

        wb.save(input_file)

        rows = merge_excel_sheets(input_file, output_file)

        assert rows == 3
        assert output_file.exists()

        # Verify merged data
        wb_out = openpyxl.load_workbook(output_file)
        ws_out = wb_out.active
        data = list(ws_out.iter_rows(values_only=True))
        assert len(data) == 4  # header + 3 rows
        assert data[0] == ("id", "name")


def test_merge_excel_sheets_specific_sheets() -> None:
    """
    Test merging specific sheets only.
    """
    with tempfile.TemporaryDirectory() as tmpdir:
        input_file = Path(tmpdir) / "input.xlsx"
        output_file = Path(tmpdir) / "output.xlsx"

        wb = openpyxl.Workbook()
        ws1 = wb.active
        ws1.title = "Data1"
        ws1.append(["id"])
        ws1.append([1])

        ws2 = wb.create_sheet("Data2")
        ws2.append(["id"])
        ws2.append([2])

        ws3 = wb.create_sheet("Ignore")
        ws3.append(["id"])
        ws3.append([999])

        wb.save(input_file)

        rows = merge_excel_sheets(
            input_file, output_file, sheet_names=["Data1", "Data2"]
        )

        assert rows == 2  # Only from Data1 and Data2


def test_merge_excel_sheets_skip_duplicates() -> None:
    """
    Test merging with duplicate row detection.
    """
    with tempfile.TemporaryDirectory() as tmpdir:
        input_file = Path(tmpdir) / "input.xlsx"
        output_file = Path(tmpdir) / "output.xlsx"

        wb = openpyxl.Workbook()
        ws1 = wb.active
        ws1.title = "Sheet1"
        ws1.append(["id", "value"])
        ws1.append([1, "A"])
        ws1.append([2, "B"])

        ws2 = wb.create_sheet("Sheet2")
        ws2.append(["id", "value"])
        ws2.append([2, "B"])  # Duplicate
        ws2.append([3, "C"])

        wb.save(input_file)

        rows = merge_excel_sheets(input_file, output_file, skip_duplicates=True)

        assert rows == 3  # Only unique rows


def test_merge_excel_sheets_no_headers() -> None:
    """
    Test merging without header rows.
    """
    with tempfile.TemporaryDirectory() as tmpdir:
        input_file = Path(tmpdir) / "input.xlsx"
        output_file = Path(tmpdir) / "output.xlsx"

        wb = openpyxl.Workbook()
        ws1 = wb.active
        ws1.append([1, "A"])
        ws1.append([2, "B"])

        ws2 = wb.create_sheet("Sheet2")
        ws2.append([3, "C"])

        wb.save(input_file)

        rows = merge_excel_sheets(input_file, output_file, include_headers=False)

        assert rows == 3


def test_merge_excel_sheets_custom_output_name() -> None:
    """
    Test merging with custom output sheet name.
    """
    with tempfile.TemporaryDirectory() as tmpdir:
        input_file = Path(tmpdir) / "input.xlsx"
        output_file = Path(tmpdir) / "output.xlsx"

        wb = openpyxl.Workbook()
        ws1 = wb.active
        ws1.append(["id"])
        ws1.append([1])

        wb.save(input_file)

        merge_excel_sheets(input_file, output_file, output_sheet_name="Combined")

        wb_out = openpyxl.load_workbook(output_file)
        assert wb_out.active.title == "Combined"


def test_merge_excel_sheets_invalid_type_input() -> None:
    """
    Test TypeError for invalid input_file type.
    """
    with pytest.raises(TypeError, match="input_file must be str or Path"):
        merge_excel_sheets(123, "output.xlsx")  # type: ignore


def test_merge_excel_sheets_invalid_type_sheet_names() -> None:
    """
    Test TypeError for invalid sheet_names type.
    """
    with tempfile.TemporaryDirectory() as tmpdir:
        input_file = Path(tmpdir) / "input.xlsx"
        output_file = Path(tmpdir) / "output.xlsx"

        with pytest.raises(TypeError, match="sheet_names must be sequence or None"):
            merge_excel_sheets(input_file, output_file, sheet_names="not_a_list")  # type: ignore


def test_merge_excel_sheets_file_not_found() -> None:
    """
    Test FileNotFoundError for non-existent file.
    """
    with tempfile.TemporaryDirectory() as tmpdir:
        output_file = Path(tmpdir) / "output.xlsx"

        with pytest.raises(FileNotFoundError, match="Input file not found"):
            merge_excel_sheets("/nonexistent/file.xlsx", output_file)


def test_merge_excel_sheets_invalid_sheet_name() -> None:
    """
    Test ValueError for non-existent sheet name.
    """
    with tempfile.TemporaryDirectory() as tmpdir:
        input_file = Path(tmpdir) / "input.xlsx"
        output_file = Path(tmpdir) / "output.xlsx"

        wb = openpyxl.Workbook()
        wb.save(input_file)

        with pytest.raises(ValueError, match="Sheet .* not found"):
            merge_excel_sheets(input_file, output_file, sheet_names=["NonExistent"])


def test_merge_excel_sheets_header_mismatch() -> None:
    """
    Test ValueError for mismatched headers.
    """
    with tempfile.TemporaryDirectory() as tmpdir:
        input_file = Path(tmpdir) / "input.xlsx"
        output_file = Path(tmpdir) / "output.xlsx"

        wb = openpyxl.Workbook()
        ws1 = wb.active
        ws1.append(["id", "name"])
        ws1.append([1, "Alice"])

        ws2 = wb.create_sheet("Sheet2")
        ws2.append(["id", "value"])  # Different header
        ws2.append([2, "B"])

        wb.save(input_file)

        with pytest.raises(ValueError, match="Header mismatch"):
            merge_excel_sheets(input_file, output_file)

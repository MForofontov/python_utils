"""Tests for read_excel_range module."""

from pathlib import Path

try:
    from openpyxl import Workbook
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    Workbook = None  # type: ignore

import pytest
from serialization_functions.excel_operations.read_excel_range import read_excel_range

pytestmark = pytest.mark.skipif(not OPENPYXL_AVAILABLE, reason="openpyxl not installed")


def test_read_excel_range_specific_range(tmp_path: Path) -> None:
    """
    Test case 1: Test reading specific cell range from Excel.
    """
    # Arrange

    file_path = tmp_path / "data.xlsx"
    wb = Workbook()
    ws = wb.active
    ws["A1"] = "Name"
    ws["B1"] = "Age"
    ws["A2"] = "Alice"
    ws["B2"] = 30
    ws["A3"] = "Bob"
    ws["B3"] = 25
    wb.save(file_path)
    wb.close()

    # Act
    data = read_excel_range(str(file_path), range_string="A1:B2")

    # Assert
    assert len(data) == 2
    assert data[0] == ["Name", "Age"]
    assert data[1] == ["Alice", 30]


def test_read_excel_range_entire_sheet(tmp_path: Path) -> None:
    """
    Test case 2: Test reading entire sheet without range specified.
    """
    # Arrange

    file_path = tmp_path / "data.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.append(["A", "B"])
    ws.append([1, 2])
    wb.save(file_path)
    wb.close()

    # Act
    data = read_excel_range(str(file_path))

    # Assert
    assert len(data) >= 2


def test_read_excel_range_by_sheet_name(tmp_path: Path) -> None:
    """
    Test case 3: Test reading from specific sheet by name.
    """
    # Arrange

    file_path = tmp_path / "multi.xlsx"
    wb = Workbook()
    ws2 = wb.create_sheet("Data")
    ws2["A1"] = "Test"
    wb.save(file_path)
    wb.close()

    # Act
    data = read_excel_range(str(file_path), sheet_name="Data", range_string="A1:A1")

    # Assert
    assert data[0][0] == "Test"


def test_read_excel_range_by_sheet_index(tmp_path: Path) -> None:
    """
    Test case 4: Test reading from specific sheet by index.
    """
    # Arrange

    file_path = tmp_path / "multi.xlsx"
    wb = Workbook()
    wb.create_sheet("Second")
    wb.worksheets[1]["A1"] = "Second Sheet"
    wb.save(file_path)
    wb.close()

    # Act
    data = read_excel_range(str(file_path), sheet_name=1, range_string="A1:A1")

    # Assert
    assert data[0][0] == "Second Sheet"


def test_read_excel_range_invalid_file_path_type_raises_error() -> None:
    """
    Test case 5: Test TypeError for invalid file_path type.
    """
    # Arrange
    expected_message = "file_path must be a string, got int"

    # Act & Assert
    with pytest.raises(TypeError, match=expected_message):
        read_excel_range(123)


def test_read_excel_range_invalid_sheet_name_type_raises_error(tmp_path: Path) -> None:
    """
    Test case 6: Test TypeError for invalid sheet_name type.
    """
    # Arrange

    file_path = tmp_path / "data.xlsx"
    wb = Workbook()
    wb.save(file_path)
    wb.close()

    expected_message = "sheet_name must be str or int, got list"

    # Act & Assert
    with pytest.raises(TypeError, match=expected_message):
        read_excel_range(str(file_path), sheet_name=["invalid"])


def test_read_excel_range_invalid_range_string_type_raises_error(
    tmp_path: Path,
) -> None:
    """
    Test case 7: Test TypeError for invalid range_string type.
    """
    # Arrange

    file_path = tmp_path / "data.xlsx"
    wb = Workbook()
    wb.save(file_path)
    wb.close()

    expected_message = "range_string must be str or None, got int"

    # Act & Assert
    with pytest.raises(TypeError, match=expected_message):
        read_excel_range(str(file_path), range_string=123)


def test_read_excel_range_invalid_sheet_index_raises_error(tmp_path: Path) -> None:
    """
    Test case 8: Test ValueError for out-of-range sheet index.
    """
    # Arrange

    file_path = tmp_path / "data.xlsx"
    wb = Workbook()
    wb.save(file_path)
    wb.close()

    # Act & Assert
    with pytest.raises(ValueError, match="sheet_name index.*out of range"):
        read_excel_range(str(file_path), sheet_name=10)


def test_read_excel_range_nonexistent_sheet_name_raises_error(tmp_path: Path) -> None:
    """
    Test case 9: Test ValueError for non-existent sheet name.
    """
    # Arrange

    file_path = tmp_path / "data.xlsx"
    wb = Workbook()
    wb.save(file_path)
    wb.close()

    # Act & Assert
    with pytest.raises(ValueError, match="sheet_name.*not found"):
        read_excel_range(str(file_path), sheet_name="NonExistent")

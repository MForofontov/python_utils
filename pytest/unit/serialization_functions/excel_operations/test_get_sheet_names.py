"""Tests for get_sheet_names module."""

from pathlib import Path

try:
    from openpyxl import Workbook
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    Workbook = None  # type: ignore

import pytest
from serialization_functions.excel_operations.get_sheet_names import get_sheet_names

pytestmark = pytest.mark.skipif(not OPENPYXL_AVAILABLE, reason="openpyxl not installed")


def test_get_sheet_names_single_sheet(tmp_path: Path) -> None:
    """
    Test case 1: Test getting sheet names from workbook with single sheet.
    """
    # Arrange

    file_path = tmp_path / "single.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Data"
    wb.save(file_path)
    wb.close()

    # Act
    sheets = get_sheet_names(str(file_path))

    # Assert
    assert isinstance(sheets, list)
    assert len(sheets) == 1
    assert sheets[0] == "Data"


def test_get_sheet_names_multiple_sheets(tmp_path: Path) -> None:
    """
    Test case 2: Test getting sheet names from workbook with multiple sheets.
    """
    # Arrange

    file_path = tmp_path / "multi.xlsx"
    wb = Workbook()
    wb.create_sheet("Sheet2")
    wb.create_sheet("Summary")
    wb.save(file_path)
    wb.close()

    # Act
    sheets = get_sheet_names(str(file_path))

    # Assert
    assert len(sheets) == 3
    assert "Sheet" in sheets
    assert "Sheet2" in sheets
    assert "Summary" in sheets


def test_get_sheet_names_preserves_order(tmp_path: Path) -> None:
    """
    Test case 3: Test that sheet names are returned in correct order.
    """
    # Arrange

    file_path = tmp_path / "ordered.xlsx"
    wb = Workbook()
    wb.active.title = "First"
    wb.create_sheet("Second")
    wb.create_sheet("Third")
    wb.save(file_path)
    wb.close()

    # Act
    sheets = get_sheet_names(str(file_path))

    # Assert
    assert sheets == ["First", "Second", "Third"]


def test_get_sheet_names_with_special_characters(tmp_path: Path) -> None:
    """
    Test case 4: Test sheet names containing special characters.
    """
    # Arrange

    file_path = tmp_path / "special.xlsx"
    wb = Workbook()
    wb.active.title = "Data-2024"
    wb.create_sheet("Summary (Q1)")
    wb.save(file_path)
    wb.close()

    # Act
    sheets = get_sheet_names(str(file_path))

    # Assert
    assert "Data-2024" in sheets
    assert "Summary (Q1)" in sheets


def test_get_sheet_names_invalid_path_type_raises_error() -> None:
    """
    Test case 6: Test TypeError for invalid file_path type.
    """
    # Arrange
    invalid_path = 123
    expected_message = "file_path must be a string, got int"

    # Act & Assert
    with pytest.raises(TypeError, match=expected_message):
        get_sheet_names(invalid_path)


def test_get_sheet_names_nonexistent_file_raises_error(tmp_path: Path) -> None:
    """
    Test case 7: Test FileNotFoundError for non-existent file.
    """
    # Arrange
    nonexistent = tmp_path / "does_not_exist.xlsx"

    # Act & Assert
    with pytest.raises(FileNotFoundError):
        get_sheet_names(str(nonexistent))


def test_get_sheet_names_invalid_file_format_raises_error(tmp_path: Path) -> None:
    """
    Test case 8: Test error for invalid Excel file format.
    """
    # Arrange
    file_path = tmp_path / "not_excel.txt"
    file_path.write_text("This is not an Excel file")

    # Act & Assert
    with pytest.raises(Exception):  # noqa: B017 - openpyxl raises various exceptions
        get_sheet_names(str(file_path))


def test_get_sheet_names_closes_workbook(tmp_path: Path) -> None:
    """
    Test case 5: Test that workbook is properly closed after reading.
    """
    # Arrange

    file_path = tmp_path / "test.xlsx"
    wb = Workbook()
    wb.save(file_path)
    wb.close()

    # Act
    sheets = get_sheet_names(str(file_path))

    # Assert - file should be accessible again
    wb2 = Workbook()
    wb2.save(file_path)  # Should not raise "file in use" error
    wb2.close()
    assert len(sheets) > 0

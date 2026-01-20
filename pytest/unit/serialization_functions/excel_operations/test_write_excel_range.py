"""Tests for write_excel_range module."""

from pathlib import Path

try:
    from openpyxl import Workbook, load_workbook
    from pyutils_collection.serialization_functions.excel_operations.write_excel_range import write_excel_range
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    Workbook = None  # type: ignore
    load_workbook = None  # type: ignore
    write_excel_range = None  # type: ignore

import pytest

pytestmark = pytest.mark.skipif(not OPENPYXL_AVAILABLE, reason="openpyxl not installed")
pytestmark = [pytestmark, pytest.mark.unit, pytest.mark.serialization]


def test_write_excel_range_basic_write(tmp_path: Path) -> None:
    """
    Test case 1: Test writing data to new Excel file.
    """
    # Arrange

    file_path = tmp_path / "output.xlsx"
    data = [["Name", "Age"], ["Alice", 30]]

    # Act
    write_excel_range(data, str(file_path))

    # Assert
    wb = load_workbook(file_path)
    ws = wb.active
    assert ws["A1"].value == "Name"
    assert ws["B1"].value == "Age"
    assert ws["A2"].value == "Alice"
    assert ws["B2"].value == 30
    wb.close()


def test_write_excel_range_custom_start_cell(tmp_path: Path) -> None:
    """
    Test case 2: Test writing data starting at custom cell.
    """
    # Arrange

    file_path = tmp_path / "output.xlsx"
    data = [["Test"]]

    # Act
    write_excel_range(data, str(file_path), start_cell="B2")

    # Assert
    wb = load_workbook(file_path)
    ws = wb.active
    assert ws["B2"].value == "Test"
    wb.close()


def test_write_excel_range_custom_sheet_name(tmp_path: Path) -> None:
    """
    Test case 3: Test writing to specific sheet name.
    """
    # Arrange

    file_path = tmp_path / "output.xlsx"
    data = [["Data"]]

    # Act
    write_excel_range(data, str(file_path), sheet_name="People")

    # Assert
    wb = load_workbook(file_path)
    assert "People" in wb.sheetnames
    wb.close()


def test_write_excel_range_append_mode(tmp_path: Path) -> None:
    """
    Test case 4: Test appending data to existing file.
    """
    # Arrange

    file_path = tmp_path / "existing.xlsx"
    wb = Workbook()
    wb.save(file_path)
    wb.close()

    data = [["New Data"]]

    # Act
    write_excel_range(data, str(file_path), mode="a")

    # Assert
    wb = load_workbook(file_path)
    ws = wb.active
    assert ws["A1"].value == "New Data"
    wb.close()


def test_write_excel_range_creates_parent_directories(tmp_path: Path) -> None:
    """
    Test case 5: Test that parent directories are created automatically.
    """
    # Arrange

    file_path = tmp_path / "subdir" / "nested" / "output.xlsx"
    data = [["Test"]]

    # Act
    write_excel_range(data, str(file_path))

    # Assert
    assert file_path.exists()
    wb = load_workbook(file_path)
    wb.close()


def test_write_excel_range_invalid_data_type_raises_error(tmp_path: Path) -> None:
    """
    Test case 6: Test TypeError for invalid data type.
    """
    # Arrange
    file_path = tmp_path / "test.xlsx"
    expected_message = "data must be a list, got str"

    # Act & Assert
    with pytest.raises(TypeError, match=expected_message):
        write_excel_range("not a list", str(file_path))


def test_write_excel_range_empty_data_raises_error(tmp_path: Path) -> None:
    """
    Test case 7: Test ValueError for empty data.
    """
    # Arrange
    file_path = tmp_path / "test.xlsx"

    # Act & Assert
    with pytest.raises(ValueError, match="data cannot be empty"):
        write_excel_range([], str(file_path))


def test_write_excel_range_invalid_mode_raises_error(tmp_path: Path) -> None:
    """
    Test case 8: Test ValueError for invalid mode.
    """
    # Arrange
    file_path = tmp_path / "test.xlsx"
    data = [["Test"]]

    # Act & Assert
    with pytest.raises(ValueError, match="mode must be 'w' or 'a'"):
        write_excel_range(data, str(file_path), mode="x")


def test_write_excel_range_invalid_start_cell_raises_error(tmp_path: Path) -> None:
    """
    Test case 9: Test ValueError for invalid start_cell format.
    """
    # Arrange
    file_path = tmp_path / "test.xlsx"
    data = [["Test"]]

    # Act & Assert
    with pytest.raises(ValueError, match="Invalid start_cell format"):
        write_excel_range(data, str(file_path), start_cell="Invalid")

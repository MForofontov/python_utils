"""Tests for create_excel_workbook module."""

from pathlib import Path

try:
    from openpyxl import load_workbook
    from python_utils.serialization_functions.excel_operations.create_excel_workbook import (
        create_excel_workbook,
    )
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    load_workbook = None  # type: ignore
    create_excel_workbook = None  # type: ignore

import pytest

pytestmark = pytest.mark.skipif(not OPENPYXL_AVAILABLE, reason="openpyxl not installed")
pytestmark = [pytestmark, pytest.mark.unit, pytest.mark.serialization]


def test_create_excel_workbook_empty(tmp_path: Path) -> None:
    """
    Test case 1: Test creating empty workbook.
    """
    # Arrange

    file_path = tmp_path / "empty.xlsx"

    # Act
    create_excel_workbook(str(file_path))

    # Assert
    assert file_path.exists()
    wb = load_workbook(file_path)
    assert len(wb.sheetnames) > 0
    wb.close()


def test_create_excel_workbook_with_data(tmp_path: Path) -> None:
    """
    Test case 2: Test creating workbook with sheet data.
    """
    # Arrange

    file_path = tmp_path / "data.xlsx"
    sheets = {"People": [["Name", "Age"], ["Alice", 30]], "Cities": [["City"], ["NYC"]]}

    # Act
    create_excel_workbook(str(file_path), sheets=sheets)

    # Assert
    wb = load_workbook(file_path)
    assert "People" in wb.sheetnames
    assert "Cities" in wb.sheetnames
    assert wb["People"]["A1"].value == "Name"
    assert wb["People"]["B2"].value == 30
    wb.close()


def test_create_excel_workbook_multiple_sheets(tmp_path: Path) -> None:
    """
    Test case 3: Test creating workbook with multiple sheets.
    """
    # Arrange

    file_path = tmp_path / "multi.xlsx"
    sheets = {"Sheet1": [["Data1"]], "Sheet2": [["Data2"]], "Sheet3": [["Data3"]]}

    # Act
    create_excel_workbook(str(file_path), sheets=sheets)

    # Assert
    wb = load_workbook(file_path)
    assert len(wb.sheetnames) == 3
    assert wb["Sheet1"]["A1"].value == "Data1"
    wb.close()


def test_create_excel_workbook_creates_parent_directories(tmp_path: Path) -> None:
    """
    Test case 4: Test that parent directories are created.
    """
    # Arrange
    file_path = tmp_path / "nested" / "folder" / "output.xlsx"

    # Act
    create_excel_workbook(str(file_path))

    # Assert
    assert file_path.exists()


def test_create_excel_workbook_invalid_file_path_type_raises_error() -> None:
    """
    Test case 5: Test TypeError for invalid file_path type.
    """
    # Arrange
    expected_message = "file_path must be a string, got int"

    # Act & Assert
    with pytest.raises(TypeError, match=expected_message):
        create_excel_workbook(123)


def test_create_excel_workbook_invalid_sheets_type_raises_error(tmp_path: Path) -> None:
    """
    Test case 6: Test TypeError for invalid sheets type.
    """
    # Arrange
    file_path = tmp_path / "test.xlsx"
    expected_message = "sheets must be a dict or None, got list"

    # Act & Assert
    with pytest.raises(TypeError, match=expected_message):
        create_excel_workbook(str(file_path), sheets=["invalid"])


def test_create_excel_workbook_invalid_sheet_name_type_raises_error(
    tmp_path: Path,
) -> None:
    """
    Test case 7: Test TypeError for invalid sheet name type.
    """
    # Arrange
    file_path = tmp_path / "test.xlsx"
    sheets = {123: [["data"]]}

    # Act & Assert
    with pytest.raises(TypeError, match="sheet name must be a string"):
        create_excel_workbook(str(file_path), sheets=sheets)


def test_create_excel_workbook_invalid_sheet_data_type_raises_error(
    tmp_path: Path,
) -> None:
    """
    Test case 8: Test TypeError for invalid sheet data type.
    """
    # Arrange
    file_path = tmp_path / "test.xlsx"
    sheets = {"Test": "not a list"}

    # Act & Assert
    with pytest.raises(TypeError, match="sheet data must be a list"):
        create_excel_workbook(str(file_path), sheets=sheets)


def test_create_excel_workbook_removes_default_sheet(tmp_path: Path) -> None:
    """
    Test case 9: Test that default sheet is removed when custom sheets provided.
    """
    # Arrange

    file_path = tmp_path / "test.xlsx"
    sheets = {"Custom": [["data"]]}

    # Act
    create_excel_workbook(str(file_path), sheets=sheets)

    # Assert
    wb = load_workbook(file_path)
    assert "Sheet" not in wb.sheetnames
    assert "Custom" in wb.sheetnames
    wb.close()

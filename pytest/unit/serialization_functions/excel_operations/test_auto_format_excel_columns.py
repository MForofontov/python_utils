"""
Unit tests for auto_format_excel_columns function.
"""

import tempfile
from pathlib import Path

try:
    import openpyxl
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    openpyxl = None  # type: ignore

import pytest
from serialization_functions.excel_operations.auto_format_excel_columns import (
    auto_format_excel_columns,
)

pytestmark = pytest.mark.skipif(not OPENPYXL_AVAILABLE, reason="openpyxl not installed")
pytestmark = [pytestmark, pytest.mark.unit, pytest.mark.serialization]


def test_auto_format_excel_columns_basic() -> None:
    """
    Test basic auto-formatting with default settings.
    """
    with tempfile.TemporaryDirectory() as tmpdir:
        file_path = Path(tmpdir) / "data.xlsx"

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Name", "Age", "Email"])
        ws.append(["Alice", 30, "alice@example.com"])
        ws.append(["Bob", 25, "bob@example.com"])

        wb.save(file_path)

        auto_format_excel_columns(file_path)

        # Verify formatting
        wb = openpyxl.load_workbook(file_path)
        ws = wb.active

        # Check frozen panes
        assert ws.freeze_panes == "A2"

        # Check bold header
        assert ws.cell(1, 1).font.bold is True


def test_auto_format_excel_columns_custom_header_color() -> None:
    """
    Test custom header background color.
    """
    with tempfile.TemporaryDirectory() as tmpdir:
        file_path = Path(tmpdir) / "data.xlsx"

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["ID", "Name"])
        ws.append([1, "Alice"])

        wb.save(file_path)
        wb.close()

        auto_format_excel_columns(file_path, header_fill_color="4472C4")

        wb = openpyxl.load_workbook(file_path)
        ws = wb.active

        # Check header fill color (openpyxl adds '00' prefix for RGB)
        assert ws.cell(1, 1).fill.start_color.rgb == "004472C4"


def test_auto_format_excel_columns_center_alignment() -> None:
    """
    Test center text alignment.
    """
    with tempfile.TemporaryDirectory() as tmpdir:
        file_path = Path(tmpdir) / "data.xlsx"

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["ID"])
        ws.append([1])

        wb.save(file_path)

        auto_format_excel_columns(file_path, align_text="center")

        wb = openpyxl.load_workbook(file_path)
        ws = wb.active

        assert ws.cell(1, 1).alignment.horizontal == "center"


def test_auto_format_excel_columns_no_freeze() -> None:
    """
    Test without freezing header row.
    """
    with tempfile.TemporaryDirectory() as tmpdir:
        file_path = Path(tmpdir) / "data.xlsx"

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["ID"])

        wb.save(file_path)

        auto_format_excel_columns(file_path, freeze_header=False)

        wb = openpyxl.load_workbook(file_path)
        ws = wb.active

        assert ws.freeze_panes is None


def test_auto_format_excel_columns_no_bold() -> None:
    """
    Test without bold header.
    """
    with tempfile.TemporaryDirectory() as tmpdir:
        file_path = Path(tmpdir) / "data.xlsx"

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["ID"])

        wb.save(file_path)

        auto_format_excel_columns(file_path, bold_header=False)

        wb = openpyxl.load_workbook(file_path)
        ws = wb.active

        # Default font is not bold
        assert ws.cell(1, 1).font.bold is not True


def test_auto_format_excel_columns_specific_sheet() -> None:
    """
    Test formatting specific sheet.
    """
    with tempfile.TemporaryDirectory() as tmpdir:
        file_path = Path(tmpdir) / "data.xlsx"

        wb = openpyxl.Workbook()
        ws1 = wb.active
        ws1.title = "Data"
        ws1.append(["ID"])

        wb.save(file_path)

        auto_format_excel_columns(file_path, sheet_name="Data")

        wb = openpyxl.load_workbook(file_path)
        ws = wb["Data"]

        assert ws.freeze_panes == "A2"


def test_auto_format_excel_columns_invalid_type_file_path() -> None:
    """
    Test TypeError for invalid file_path type.
    """
    with pytest.raises(TypeError, match="file_path must be str or Path"):
        auto_format_excel_columns(123)  # type: ignore


def test_auto_format_excel_columns_invalid_align_text() -> None:
    """
    Test ValueError for invalid align_text value.
    """
    with tempfile.TemporaryDirectory() as tmpdir:
        file_path = Path(tmpdir) / "data.xlsx"

        wb = openpyxl.Workbook()
        wb.save(file_path)

        with pytest.raises(ValueError, match="align_text must be"):
            auto_format_excel_columns(file_path, align_text="invalid")


def test_auto_format_excel_columns_file_not_found() -> None:
    """
    Test FileNotFoundError for non-existent file.
    """
    with pytest.raises(FileNotFoundError, match="Excel file not found"):
        auto_format_excel_columns("/nonexistent/file.xlsx")


def test_auto_format_excel_columns_invalid_sheet_name() -> None:
    """
    Test ValueError for non-existent sheet name.
    """
    with tempfile.TemporaryDirectory() as tmpdir:
        file_path = Path(tmpdir) / "data.xlsx"

        wb = openpyxl.Workbook()
        wb.save(file_path)

        with pytest.raises(ValueError, match="Sheet .* not found"):
            auto_format_excel_columns(file_path, sheet_name="NonExistent")

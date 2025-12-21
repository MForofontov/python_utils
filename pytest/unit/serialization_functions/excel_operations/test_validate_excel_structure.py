"""
Unit tests for validate_excel_structure function.
"""

from pathlib import Path
import tempfile

import pytest
import openpyxl
from serialization_functions.excel_operations.validate_excel_structure import (
    validate_excel_structure,
)


def test_validate_excel_structure_valid() -> None:
    """
    Test validation of valid Excel structure.
    """
    with tempfile.TemporaryDirectory() as tmpdir:
        excel_file = Path(tmpdir) / "data.xlsx"

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["ID", "Name", "Email"])
        ws.append([1, "Alice", "alice@example.com"])
        ws.append([2, "Bob", "bob@example.com"])

        wb.save(excel_file)
        wb.close()

        result = validate_excel_structure(excel_file)

        assert result["valid"] is True
        assert result["row_count"] == 2
        assert result["columns"] == ["ID", "Name", "Email"]
        assert len(result["errors"]) == 0


def test_validate_excel_structure_required_columns() -> None:
    """
    Test validation with required columns.
    """
    with tempfile.TemporaryDirectory() as tmpdir:
        excel_file = Path(tmpdir) / "data.xlsx"

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["ID", "Name"])
        ws.append([1, "Alice"])

        wb.save(excel_file)
        wb.close()

        result = validate_excel_structure(excel_file, required_columns=["ID", "Name"])

        assert result["valid"] is True


def test_validate_excel_structure_missing_columns() -> None:
    """
    Test detection of missing required columns.
    """
    with tempfile.TemporaryDirectory() as tmpdir:
        excel_file = Path(tmpdir) / "data.xlsx"

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["ID", "Name"])
        ws.append([1, "Alice"])

        wb.save(excel_file)
        wb.close()

        result = validate_excel_structure(
            excel_file, required_columns=["ID", "Name", "Email"]
        )

        assert result["valid"] is False
        assert any("Missing required columns" in error for error in result["errors"])


def test_validate_excel_structure_row_count_limits() -> None:
    """
    Test row count validation.
    """
    with tempfile.TemporaryDirectory() as tmpdir:
        excel_file = Path(tmpdir) / "data.xlsx"

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["ID"])
        for i in range(5):
            ws.append([i])

        wb.save(excel_file)
        wb.close()

        result = validate_excel_structure(excel_file, min_rows=3, max_rows=10)

        assert result["valid"] is True
        assert result["row_count"] == 5


def test_validate_excel_structure_too_few_rows() -> None:
    """
    Test error for insufficient rows.
    """
    with tempfile.TemporaryDirectory() as tmpdir:
        excel_file = Path(tmpdir) / "data.xlsx"

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["ID"])
        ws.append([1])

        wb.save(excel_file)
        wb.close()

        result = validate_excel_structure(excel_file, min_rows=5)

        assert result["valid"] is False
        assert any("Insufficient rows" in error for error in result["errors"])


def test_validate_excel_structure_too_many_rows() -> None:
    """
    Test error for exceeding max rows.
    """
    with tempfile.TemporaryDirectory() as tmpdir:
        excel_file = Path(tmpdir) / "data.xlsx"

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["ID"])
        for i in range(20):
            ws.append([i])

        wb.save(excel_file)
        wb.close()

        result = validate_excel_structure(excel_file, max_rows=10)

        assert result["valid"] is False
        assert any("Too many rows" in error for error in result["errors"])


def test_validate_excel_structure_strict_columns() -> None:
    """
    Test strict column matching.
    """
    with tempfile.TemporaryDirectory() as tmpdir:
        excel_file = Path(tmpdir) / "data.xlsx"

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["ID", "Name", "Extra"])
        ws.append([1, "Alice", "X"])

        wb.save(excel_file)
        wb.close()

        result = validate_excel_structure(
            excel_file, required_columns=["ID", "Name"], strict_columns=True
        )

        assert result["valid"] is False
        assert any("Extra columns" in error for error in result["errors"])


def test_validate_excel_structure_duplicate_columns() -> None:
    """
    Test warning for duplicate column names.
    """
    with tempfile.TemporaryDirectory() as tmpdir:
        excel_file = Path(tmpdir) / "data.xlsx"

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["ID", "Name", "Name"])
        ws.append([1, "Alice", "Bob"])

        wb.save(excel_file)
        wb.close()

        result = validate_excel_structure(excel_file)

        assert any("Duplicate column name" in warning for warning in result["warnings"])


def test_validate_excel_structure_empty_header() -> None:
    """
    Test warning for empty column headers.
    """
    with tempfile.TemporaryDirectory() as tmpdir:
        excel_file = Path(tmpdir) / "data.xlsx"

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["ID", None, "Name"])
        ws.append([1, "X", "Alice"])

        wb.save(excel_file)
        wb.close()

        result = validate_excel_structure(excel_file)

        assert any("Empty column header" in warning for warning in result["warnings"])


def test_validate_excel_structure_specific_sheet() -> None:
    """
    Test validation of specific sheet.
    """
    with tempfile.TemporaryDirectory() as tmpdir:
        excel_file = Path(tmpdir) / "data.xlsx"

        wb = openpyxl.Workbook()
        ws1 = wb.active
        ws1.title = "Sheet1"
        ws1.append(["Data"])

        ws2 = wb.create_sheet("Sheet2")
        ws2.append(["ID"])
        ws2.append([1])

        wb.save(excel_file)
        wb.close()

        result = validate_excel_structure(excel_file, sheet_name="Sheet2")

        assert result["valid"] is True
        assert result["columns"] == ["ID"]


def test_validate_excel_structure_invalid_type_file_path() -> None:
    """
    Test TypeError for invalid file_path type.
    """
    with pytest.raises(TypeError, match="file_path must be str or Path"):
        validate_excel_structure(123)  # type: ignore


def test_validate_excel_structure_invalid_min_rows() -> None:
    """
    Test ValueError for negative min_rows.
    """
    with tempfile.TemporaryDirectory() as tmpdir:
        excel_file = Path(tmpdir) / "data.xlsx"

        wb = openpyxl.Workbook()
        wb.save(excel_file)
        wb.close()

        with pytest.raises(ValueError, match="min_rows must be non-negative"):
            validate_excel_structure(excel_file, min_rows=-1)


def test_validate_excel_structure_file_not_found() -> None:
    """
    Test FileNotFoundError for non-existent file.
    """
    with pytest.raises(FileNotFoundError, match="Excel file not found"):
        validate_excel_structure("/nonexistent/file.xlsx")


def test_validate_excel_structure_invalid_sheet_name() -> None:
    """
    Test error for non-existent sheet.
    """
    with tempfile.TemporaryDirectory() as tmpdir:
        excel_file = Path(tmpdir) / "data.xlsx"

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        wb.save(excel_file)
        wb.close()

        result = validate_excel_structure(excel_file, sheet_name="NonExistent")

        assert result["valid"] is False
        assert any("not found" in error for error in result["errors"])

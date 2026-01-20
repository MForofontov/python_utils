"""
Unit tests for transpose_excel_data function.
"""

import tempfile
from pathlib import Path

try:
    import openpyxl
    from python_utils.serialization_functions.excel_operations.transpose_excel_data import (
        transpose_excel_data,
    )
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    openpyxl = None  # type: ignore
    transpose_excel_data = None  # type: ignore

import pytest

pytestmark = pytest.mark.skipif(not OPENPYXL_AVAILABLE, reason="openpyxl not installed")
pytestmark = [pytestmark, pytest.mark.unit, pytest.mark.serialization]


def test_transpose_excel_data_basic() -> None:
    """
    Test basic transpose operation.
    """
    with tempfile.TemporaryDirectory() as tmpdir:
        input_file = Path(tmpdir) / "input.xlsx"
        output_file = Path(tmpdir) / "output.xlsx"

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Name", "Age", "City"])
        ws.append(["Alice", 30, "NYC"])
        ws.append(["Bob", 25, "LA"])

        wb.save(input_file)

        rows, cols = transpose_excel_data(input_file, output_file)

        assert rows == 3
        assert cols == 3
        assert output_file.exists()

        # Verify transposed data
        wb_out = openpyxl.load_workbook(output_file)
        ws_out = wb_out.active
        data = list(ws_out.iter_rows(values_only=True))
        assert len(data) == 3  # Original columns become rows
        assert data[0] == ("Name", "Alice", "Bob")
        assert data[1] == ("Age", 30, 25)


def test_transpose_excel_data_specific_sheet() -> None:
    """
    Test transposing specific sheet.
    """
    with tempfile.TemporaryDirectory() as tmpdir:
        input_file = Path(tmpdir) / "input.xlsx"
        output_file = Path(tmpdir) / "output.xlsx"

        wb = openpyxl.Workbook()
        ws1 = wb.active
        ws1.title = "Data"
        ws1.append([1, 2, 3])
        ws1.append([4, 5, 6])

        wb.save(input_file)

        rows, cols = transpose_excel_data(
            input_file, output_file, sheet_name="Data", output_sheet_name="Transposed"
        )

        assert rows == 2
        assert cols == 3

        wb_out = openpyxl.load_workbook(output_file)
        assert wb_out.active.title == "Transposed"


def test_transpose_excel_data_empty_sheet() -> None:
    """
    Test transposing empty sheet.
    """
    with tempfile.TemporaryDirectory() as tmpdir:
        input_file = Path(tmpdir) / "input.xlsx"
        output_file = Path(tmpdir) / "output.xlsx"

        wb = openpyxl.Workbook()
        wb.save(input_file)

        rows, cols = transpose_excel_data(input_file, output_file)

        assert rows == 0
        assert cols == 0


def test_transpose_excel_data_single_row() -> None:
    """
    Test transposing single row.
    """
    with tempfile.TemporaryDirectory() as tmpdir:
        input_file = Path(tmpdir) / "input.xlsx"
        output_file = Path(tmpdir) / "output.xlsx"

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append([1, 2, 3, 4, 5])

        wb.save(input_file)

        rows, cols = transpose_excel_data(input_file, output_file)

        assert rows == 1
        assert cols == 5

        # Should become 5 rows, 1 column
        wb_out = openpyxl.load_workbook(output_file)
        ws_out = wb_out.active
        data = list(ws_out.iter_rows(values_only=True))
        assert len(data) == 5


def test_transpose_excel_data_irregular_rows() -> None:
    """
    Test transposing irregular row lengths.
    """
    with tempfile.TemporaryDirectory() as tmpdir:
        input_file = Path(tmpdir) / "input.xlsx"
        output_file = Path(tmpdir) / "output.xlsx"

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append([1, 2, 3])
        ws.append([4, 5])  # Shorter row
        ws.append([6, 7, 8, 9])  # Longer row

        wb.save(input_file)

        rows, cols = transpose_excel_data(input_file, output_file)

        assert rows == 3
        assert cols == 4  # Max column length


def test_transpose_excel_data_invalid_type_input() -> None:
    """
    Test TypeError for invalid input_file type.
    """
    with pytest.raises(TypeError, match="input_file must be str or Path"):
        transpose_excel_data(123, "output.xlsx")  # type: ignore


def test_transpose_excel_data_invalid_type_sheet_name() -> None:
    """
    Test TypeError for invalid sheet_name type.
    """
    with tempfile.TemporaryDirectory() as tmpdir:
        input_file = Path(tmpdir) / "input.xlsx"
        output_file = Path(tmpdir) / "output.xlsx"

        with pytest.raises(TypeError, match="sheet_name must be str or None"):
            transpose_excel_data(input_file, output_file, sheet_name=123)  # type: ignore


def test_transpose_excel_data_file_not_found() -> None:
    """
    Test FileNotFoundError for non-existent file.
    """
    with tempfile.TemporaryDirectory() as tmpdir:
        output_file = Path(tmpdir) / "output.xlsx"

        with pytest.raises(FileNotFoundError, match="Input file not found"):
            transpose_excel_data("/nonexistent/file.xlsx", output_file)


def test_transpose_excel_data_invalid_sheet_name() -> None:
    """
    Test ValueError for non-existent sheet name.
    """
    with tempfile.TemporaryDirectory() as tmpdir:
        input_file = Path(tmpdir) / "input.xlsx"
        output_file = Path(tmpdir) / "output.xlsx"

        wb = openpyxl.Workbook()
        wb.save(input_file)

        with pytest.raises(ValueError, match="Sheet .* not found"):
            transpose_excel_data(input_file, output_file, sheet_name="NonExistent")

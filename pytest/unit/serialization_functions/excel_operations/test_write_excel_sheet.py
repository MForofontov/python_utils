"""Tests for write_excel_sheet module."""

from pathlib import Path

try:
    from openpyxl import Workbook, load_workbook
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    Workbook = None  # type: ignore
    load_workbook = None  # type: ignore

import pytest
from python_utils.serialization_functions.excel_operations.write_excel_sheet import write_excel_sheet

pytestmark = pytest.mark.skipif(not OPENPYXL_AVAILABLE, reason="openpyxl not installed")
pytestmark = [pytestmark, pytest.mark.unit, pytest.mark.serialization]


def test_write_excel_sheet_basic_write(tmp_path: Path) -> None:
    """
    Test case 1: Test basic writing to Excel sheet.
    """
    # Arrange

    file_path = tmp_path / "output.xlsx"
    data = [["Name", "Age"], ["Alice", 30]]

    # Act
    write_excel_sheet(data, str(file_path))

    # Assert
    wb = load_workbook(file_path)
    ws = wb.active
    assert ws.cell(1, 1).value == "Name"
    assert ws.cell(2, 2).value == 30
    wb.close()


def test_write_excel_sheet_custom_start_position(tmp_path: Path) -> None:
    """
    Test case 2: Test writing with custom start row and column.
    """
    # Arrange

    file_path = tmp_path / "output.xlsx"
    data = [["Test"]]

    # Act
    write_excel_sheet(data, str(file_path), start_row=3, start_col=2)

    # Assert
    wb = load_workbook(file_path)
    ws = wb.active
    assert ws.cell(3, 2).value == "Test"
    wb.close()


def test_write_excel_sheet_custom_sheet_name(tmp_path: Path) -> None:
    """
    Test case 3: Test writing to specific sheet name.
    """
    # Arrange

    file_path = tmp_path / "output.xlsx"
    data = [["Data"]]

    # Act
    write_excel_sheet(data, str(file_path), sheet_name="CustomSheet")

    # Assert
    wb = load_workbook(file_path)
    assert "CustomSheet" in wb.sheetnames
    wb.close()


def test_write_excel_sheet_append_mode(tmp_path: Path) -> None:
    """
    Test case 4: Test appending to existing file.
    """
    # Arrange

    file_path = tmp_path / "existing.xlsx"
    wb = Workbook()
    wb.save(file_path)
    wb.close()

    data = [["New"]]

    # Act
    write_excel_sheet(data, str(file_path), mode="a")

    # Assert
    wb = load_workbook(file_path)
    ws = wb.active
    assert ws.cell(1, 1).value == "New"
    wb.close()


def test_write_excel_sheet_invalid_data_type_raises_error(tmp_path: Path) -> None:
    """
    Test case 5: Test TypeError for invalid data type.
    """
    # Arrange
    file_path = tmp_path / "test.xlsx"

    # Act & Assert
    with pytest.raises(TypeError, match="data must be a list"):
        write_excel_sheet("invalid", str(file_path))


def test_write_excel_sheet_empty_data_raises_error(tmp_path: Path) -> None:
    """
    Test case 6: Test ValueError for empty data.
    """
    # Arrange
    file_path = tmp_path / "test.xlsx"

    # Act & Assert
    with pytest.raises(ValueError, match="data cannot be empty"):
        write_excel_sheet([], str(file_path))


def test_write_excel_sheet_invalid_start_row_raises_error(tmp_path: Path) -> None:
    """
    Test case 7: Test ValueError for invalid start_row.
    """
    # Arrange
    file_path = tmp_path / "test.xlsx"
    data = [["Test"]]

    # Act & Assert
    with pytest.raises(ValueError, match="start_row must be >= 1"):
        write_excel_sheet(data, str(file_path), start_row=0)


def test_write_excel_sheet_invalid_start_col_raises_error(tmp_path: Path) -> None:
    """
    Test case 8: Test ValueError for invalid start_col.
    """
    # Arrange
    file_path = tmp_path / "test.xlsx"
    data = [["Test"]]

    # Act & Assert
    with pytest.raises(ValueError, match="start_col must be >= 1"):
        write_excel_sheet(data, str(file_path), start_col=-1)


def test_write_excel_sheet_creates_parent_directories(tmp_path: Path) -> None:
    """
    Test case 9: Test that parent directories are created.
    """
    # Arrange
    file_path = tmp_path / "nested" / "output.xlsx"
    data = [["Test"]]

    # Act
    write_excel_sheet(data, str(file_path))

    # Assert
    assert file_path.exists()
